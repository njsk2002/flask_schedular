from datetime import datetime, timedelta
from openpyxl import Workbook
import openpyxl    
from flask import Blueprint, render_template, request, url_for, g , flash, jsonify
from werkzeug.utils import redirect
from pybo import db  # .. import db로 되어있는데, pybo로 변경
from ..forms import QuestionForm, AnswerForm
from sqlalchemy import func
from ..models import Question, Answer, User, QuestionVoter, WaterUsage,ElecUsage,VehicleUsage,Co2Management,Stockinfo, StockListFromVB

from pybo.views.auth_views import login_required
from ..service.co2_cal import Co2Cal

bp = Blueprint('co2',__name__,url_prefix='/co2')


@bp.route('/mainlist', methods=['GET', 'POST'])
def main_list():
    main_list = Co2Management.query.order_by(Co2Management.no).all()

     # 합계 계산
    total_elec = sum(float(item.co2_elec) if item.co2_elec is not None else 0.0 for item in main_list)
    total_water = sum(float(item.co2_water) if item.co2_water is not None else 0.0 for item in main_list)
    total_vehicle = sum(float(item.co2_vehicle) if item.co2_vehicle is not None else 0.0 for item in main_list)
    total_waste = sum(float(item.co2_waste) if item.co2_waste is not None else 0.0 for item in main_list)
    total_gas = sum(float(item.co2_gas) if item.co2_gas is not None else 0.0 for item in main_list)
    total_all = sum(float(item.use_total) if item.use_total is not None else 0.0 for item in main_list)

    total_use_elec = sum(float(item.use_elec) if item.use_elec is not None else 0.0 for item in main_list)
    total_use_water = sum(float(item.use_water) if item.use_water is not None else 0.0 for item in main_list)
    total_use_vehicle = sum(float(item.use_vehicle) if item.use_vehicle is not None else 0.0 for item in main_list)
    total_use_waste = sum(float(item.use_waste) if item.use_waste is not None else 0.0 for item in main_list)
    total_use_gas = sum(float(item.use_gas) if item.use_gas is not None else 0.0 for item in main_list)
    total_use_all = total_use_elec + total_use_water + total_use_vehicle + total_use_gas


    # 각 필드에서 값이 0이 아닌 항목의 수 계산
    count_elec = sum(1 for item in main_list if item.co2_elec not in [None, 0.0])
    count_water = sum(1 for item in main_list if item.co2_water not in [None, 0.0])
    count_vehicle = sum(1 for item in main_list if item.co2_vehicle not in [None, 0.0])
    count_waste = sum(1 for item in main_list if item.co2_waste not in [None, 0.0])
    count_gas = sum(1 for item in main_list if item.co2_gas not in [None, 0.0])
    count_total = sum(1 for item in main_list if item.use_total not in [None, 0.0])

    # co2_elec 기준으로 0이 아닌 마지막 달 찾기
    last_valid_date = None
    for item in reversed(main_list):
        if item.co2_elec not in [None, 0.0,'0.0',0]:
            last_valid_date = item.use_date
            break
    #print(last_valid_date)
    # 마지막 유효 날짜 이후의 데이터는 포함하지 않도록 필터링
    #filtered_list = [item for item in main_list if item.use_date == last_valid_date]

    # filtered_list를 사전 형식으로 변환
    filtered_list = [
    {
        'use_date': item.use_date,  # 날짜는 그대로
        'co2_elec': float(item.co2_elec) if item.co2_elec is not None else 0.0,
        'co2_water': float(item.co2_water) if item.co2_water is not None else 0.0,
        'co2_vehicle': float(item.co2_vehicle) if item.co2_vehicle is not None else 0.0,
        'co2_waste': float(item.co2_waste) if item.co2_waste is not None else 0.0,
        'co2_gas': float(item.co2_gas) if item.co2_gas is not None else 0.0,
        'use_elec': float(item.use_elec) if item.use_elec is not None else 0.0,
        'use_water': float(item.use_water) if item.use_water is not None else 0.0,
        'use_vehicle': float(item.use_vehicle) if item.use_vehicle is not None else 0.0,
        'use_waste': float(item.use_waste) if item.use_waste is not None else 0.0,
        'use_gas': float(item.use_gas) if item.use_gas is not None else 0.0,
        'use_total': float(item.use_total) if item.use_total is not None else 0.0
    }
    for item in main_list if item.use_date == last_valid_date
]

    # for item in filtered_list:
    #     print(f"Date: {item.use_date}, CO2 Electricity: {item.co2_elec}, CO2 Water: {item.co2_water}, CO2 Vehicle: {item.co2_vehicle}, CO2 Waste: {item.co2_waste}, CO2 Gas: {item.co2_gas}, Total: {item.use_total}")

    # Co2평균 계산 (각 필드에서 값이 0이 아닌 항목의 수로 나누기)
    avg_elec = total_elec / count_elec if count_elec > 0 else 0
    avg_water = total_water / count_water if count_water > 0 else 0
    avg_vehicle = total_vehicle / count_vehicle if count_vehicle > 0 else 0
    avg_waste = total_waste / count_waste if count_waste > 0 else 0
    avg_gas = total_gas / count_gas if count_gas > 0 else 0
    avg_total = total_all / count_total if count_total > 0 else 0

    # 사용량 평균 계산 (각 필드에서 값이 0이 아닌 항목의 수로 나누기)
    avg_use_elec = total_use_elec / count_elec if count_elec > 0 else 0
    avg_use_water = total_use_water / count_water if count_water > 0 else 0
    avg_use_vehicle = total_use_vehicle / count_vehicle if count_vehicle > 0 else 0
    avg_use_waste = total_use_waste / count_waste if count_waste > 0 else 0
    avg_use_gas = total_use_gas / count_gas if count_gas > 0 else 0
    avg_use_total_use = total_all / count_total if count_total > 0 else 0

    sum_data = {
        'main_list': main_list,
        'totals': {
            'elec': round(total_elec, 2),
            'water': round(total_water, 2),
            'vehicle': round(total_vehicle, 2),
            'waste': round(total_waste, 2),
            'gas': round(total_gas, 2),
            'total': round(total_all, 2),
            'use_elec': round(total_use_elec, 2),
            'use_water': round(total_use_water, 2),
            'use_vehicle': round(total_use_vehicle, 2),
            'use_waste': round(total_use_waste, 2),
            'use_gas': round(total_use_gas, 2),
            'use_total': round(total_use_all, 2)
        },
        'averages': {
            'elec': round(avg_elec, 2),
            'water': round(avg_water, 2),
            'vehicle': round(avg_vehicle, 2),
            'waste': round(avg_waste, 2),
            'gas': round(avg_gas, 2),
            'total': round(avg_total, 2),
            'use_elec': round(avg_use_elec, 2),
            'use_water': round(avg_use_water, 2),
            'use_vehicle': round(avg_use_vehicle, 2),
            'use_waste': round(avg_use_waste, 2),
            'use_gas': round(avg_use_gas, 2),
            'use_total': round(avg_use_total_use, 2)
        }
    }

    #print(sum_data)
    data = {
                'labels': [f"{item.use_date}" for item in main_list],
                'co2_elec': [round(float(item.co2_elec), 2) if item.co2_elec is not None else 0.0 for item in main_list],
                'co2_water': [round(float(item.co2_water), 2) if item.co2_water is not None else 0.0 for item in main_list],
                'co2_vehicle': [round(float(item.co2_vehicle), 2) if item.co2_vehicle is not None else 0.0 for item in main_list],
                'co2_waste': [round(float(item.co2_waste), 2) if item.co2_waste is not None else 0.0 for item in main_list],
                'co2_gas': [round(float(item.co2_gas), 2) if item.co2_gas is not None else 0.0 for item in main_list],
                'current_data': filtered_list  # 문제 없이 사용 가능
            }

    print(data)
   
    # if request.method == 'POST':
    #     #return jsonify(data)  # For POST requests, return the data as JSON
    #     return render_template('co2/main_co2.html', data=data, sum_data = sum_data)  # For GET requests, render the template with data
    # else:
    return render_template('co2/main_co2.html', data=data, sum_data = sum_data)  # For GET requests, render the template with data


@bp.route('/details/', methods=['GET'])
def _details():
    #입력파라메터
    page_co2 = request.args.get('page_co2', type=int, default=1)
    page_water = request.args.get('page_water', type=int, default=1)
    page_elec = request.args.get('page_elec', type=int, default=1)
    page_vehicle = request.args.get('page_vehicle', type=int, default=1)
    kw = request.args.get('kw',type=str, default='')
    so = request.args.get('so', type=str, default='recent')
    
    #정렬
    if so == 'recommend':
        sub_query = db.session.query(QuestionVoter.question_id, func.count('*').label('num_voter'))\
                                     .group_by(QuestionVoter.question_id).subquery()

        question_list = Question.query.outerjoin(sub_query, Question.id == sub_query.c.question_id)\
        .order_by(sub_query.c.num_voter.desc(), Question.create_date.desc())
    elif so == 'popular':
        sub_query = db.session.query(Answer.question_id, func.count('*').label('num_answer'))\
                                     .group_by(Answer.question_id).subquery()

        question_list = Question.query.outerjoin(sub_query, Question.id == sub_query.c.question_id)\
        .order_by(sub_query.c.num_answer.desc(), Question.create_date.desc())
    else:
        co2_list = Co2Management.query.order_by( Co2Management.no)
        water_use = WaterUsage.query.order_by(WaterUsage.no)
        elec_use = ElecUsage.query.order_by(ElecUsage.no)
        vehicle_use = VehicleUsage.query.order_by(VehicleUsage.no)
    
        # CO2 관리 데이터를 가져올 때 None 값을 0으로 처리
        # co2_list = [
        #     {
        #         'no': item.no,
        #         'co2_elec': item.co2_elec if item.co2_elec is not None else 0.0,
        #         'co2_water': item.co2_water if item.co2_water is not None else 0.0,
        #         'co2_vehicle': item.co2_vehicle if item.co2_vehicle is not None else 0.0,
        #         'co2_waste': item.co2_waste if item.co2_waste is not None else 0.0,
        #         'co2_gas': item.co2_gas if item.co2_gas is not None else 0.0,
        #     }
        #     for item in Co2Management.query.order_by(Co2Management.no)
        # ]

    
    # #조회
    # question_list =  Question.query.order_by(Question.create_date.desc()) 
    # 조회
    if kw:
        search = '%%{}%%'.format(kw)
        sub_query = db.session.query(Answer.question_id, Answer.content, User.username) \
            .join(User, Answer.user_no == User.no).subquery()
        question_list = question_list \
            .join(User) \
            .outerjoin(sub_query, sub_query.c.question_id == Question.id) \
            .filter(Question.subject.ilike(search) |  # 질문제목
                    Question.content.ilike(search) |  # 질문내용
                    User.username.ilike(search) |  # 질문작성자
                    sub_query.c.content.ilike(search) |  # 답변내용
                    sub_query.c.username.ilike(search)  # 답변작성자
                    ) \
            .distinct()

    # 페이징
    co2_list = co2_list.paginate(page=page_co2, per_page=10)
    water_use = water_use.paginate(page=page_water, per_page=10)
    elec_use = elec_use.paginate(page=page_elec, per_page=10)
    vehicle_use = vehicle_use.paginate(page=page_vehicle, per_page=10)

    return render_template('co2/co2_details.html', 
                            co2_list=co2_list,
                            water_use=water_use,
                            elec_use=elec_use,
                            vehicle_use=vehicle_use,
                            page_co2=page_co2, 
                            page_water=page_water, 
                            page_elec=page_elec, 
                            page_vehicle=page_vehicle, 
                            kw=kw)

@bp.route('/admin')
@login_required
def admin():  
    #입력파라메터
    page_co2 = request.args.get('page_co2', type=int, default=1)
    page_water = request.args.get('page_water', type=int, default=1)
    page_elec = request.args.get('page_elec', type=int, default=1)
    page_vehicle = request.args.get('page_vehicle', type=int, default=1)
    kw = request.args.get('kw',type=str, default='')
    so = request.args.get('so', type=str, default='recent')

    current_date = datetime.now()
    selected_year = current_date.year
    selected_month = current_date.month
    
    #정렬
    if so == 'recommend':
        sub_query = db.session.query(QuestionVoter.question_id, func.count('*').label('num_voter'))\
                                     .group_by(QuestionVoter.question_id).subquery()

        question_list = Question.query.outerjoin(sub_query, Question.id == sub_query.c.question_id)\
        .order_by(sub_query.c.num_voter.desc(), Question.create_date.desc())
    elif so == 'popular':
        sub_query = db.session.query(Answer.question_id, func.count('*').label('num_answer'))\
                                     .group_by(Answer.question_id).subquery()

        question_list = Question.query.outerjoin(sub_query, Question.id == sub_query.c.question_id)\
        .order_by(sub_query.c.num_answer.desc(), Question.create_date.desc())
    else:
        co2_list = Co2Management.query.order_by( Co2Management.no)
        water_use = WaterUsage.query.order_by(WaterUsage.no)
        elec_use = ElecUsage.query.order_by(ElecUsage.no)
        vehicle_use = VehicleUsage.query.order_by(VehicleUsage.no)

 
    # 조회
    if kw:
        search = f"%{kw}%"
        co2_list = co2_list.filter(Co2Management.subject.ilike(search))
        water_use = water_use.filter(WaterUsage.subject.ilike(search))
        elec_use = elec_use.filter(ElecUsage.subject.ilike(search))
        vehicle_use = vehicle_use.filter(VehicleUsage.subject.ilike(search))

    #db 업데이트를 위한 함수 호출
    
    # 사용량 업데이트
    Co2Cal.update_elec()  
    Co2Cal.update_water()  
    Co2Cal.update_vehicle() 
    Co2Cal.sum_update()

    # 페이징
    co2_list = co2_list.paginate(page=page_co2, per_page=10)
    water_use = water_use.paginate(page=page_water, per_page=10)
    elec_use = elec_use.paginate(page=page_elec, per_page=10)
    vehicle_use = vehicle_use.paginate(page=page_vehicle, per_page=10) 
    
    return render_template('co2/admin.html',
                            co2_list=co2_list,
                            water_use=water_use,
                            elec_use=elec_use,
                            vehicle_use=vehicle_use,
                            page_co2=page_co2, 
                            page_water=page_water, 
                            page_elec=page_elec, 
                            page_vehicle=page_vehicle, 
                            kw=kw,
                            current_date=current_date, 
                            selected_year=selected_year,
                            selected_month=selected_month                       
                            )

################### Admin 전기사용량 리딩 ###################################################
@bp.route('/get_elec_data')
def get_elec_data():
    year = request.args.get('year')
    month = int(request.args.get('month'))
    prev_year = request.args.get('prev_year')
    prev_month = int(request.args.get('prev_month'))

    month_map = {
        1: "jan",
        2: "feb",
        3: "mar",
        4: "apr",
        5: "may",
        6: "jun",
        7: "jul",
        8: "aug",
        9: "sep",
        10: "oct",
        11: "nov",
        12: "dec"
    }

    month_column = f'acum_{month_map[month]}'  # e.g., acum_jul
    prev_month_column = f'acum_{month_map[prev_month]}'  # e.g., acum_jun

    elec_data = {}
    for room_no in range(201, 211):
        current_record = ElecUsage.query.filter_by(use_year=year, room_no=f'{room_no}').first()
        prev_record = ElecUsage.query.filter_by(use_year=prev_year, room_no=f'{room_no}').first()

        elec_data[f'elec_current_{room_no}'] = getattr(current_record, month_column, 0)
        elec_data[f'elec_prev_{room_no}'] = getattr(prev_record, prev_month_column, 0)
    
    print(elec_data)
    return jsonify(elec_data)


################### Admin 전기 사용량 입력 #####################################################
@bp.route('/save_elec_data', methods=['POST'])
def save_elec_data():
    data = request.json
    selected_year = data.get('selected_year')
    selected_month = int(data.get('selected_month'))

    month_map = {
        1: "jan",
        2: "feb",
        3: "mar",
        4: "apr",
        5: "may",
        6: "jun",
        7: "jul",
        8: "aug",
        9: "sep",
        10: "oct",
        11: "nov",
        12: "dec"
    }

    # month_column = f'acum_{month_map[selected_month]}'  # e.g., acum_jul
    

    for room_no in range(201, 211):
        current_value = data.get(f'elec_current_{room_no}')
        if current_value is not None:
            # Update the database record for the specific room and month
            ElecUsage.query.filter_by(use_year=selected_year, room_no=f'{room_no}').update({
                f'acum_{month_map[selected_month]}': current_value,
            })

    db.session.commit()
    return jsonify({'status': 'success'})


################### Admin 수도사용량 리딩 ###################################################
@bp.route('/get_water_data')
def get_water_data():
    year = request.args.get('year')
    month = int(request.args.get('month'))
    prev_year = request.args.get('prev_year')
    prev_month = int(request.args.get('prev_month'))

    month_map = {
        1: "jan",
        2: "feb",
        3: "mar",
        4: "apr",
        5: "may",
        6: "jun",
        7: "jul",
        8: "aug",
        9: "sep",
        10: "oct",
        11: "nov",
        12: "dec"
    }

    month_column = f'acum_{month_map[month]}'  # e.g., acum_jul
    prev_month_column = f'acum_{month_map[prev_month]}'  # e.g., acum_jun

    water_data = {}
    for room_no in range(201, 211):
        current_record =WaterUsage.query.filter_by(use_year=year, room_no=f'{room_no}').first()
        prev_record = WaterUsage.query.filter_by(use_year=prev_year, room_no=f'{room_no}').first()

        water_data[f'water_current_{room_no}'] = getattr(current_record, month_column, 0)
        water_data[f'water_prev_{room_no}'] = getattr(prev_record, prev_month_column, 0)
    
    print(water_data)
    return jsonify(water_data)


################### Admin 수도 사용량 입력 #####################################################
@bp.route('/save_water_data', methods=['POST'])
def save_water_data():
    data = request.json
    selected_year = data.get('selected_year')
    selected_month = int(data.get('selected_month'))

    month_map = {
        1: "jan",
        2: "feb",
        3: "mar",
        4: "apr",
        5: "may",
        6: "jun",
        7: "jul",
        8: "aug",
        9: "sep",
        10: "oct",
        11: "nov",
        12: "dec"
    }

    # month_column = f'acum_{month_map[selected_month]}'  # e.g., acum_jul
    

    for room_no in range(201, 211):
        current_value = data.get(f'water_current_{room_no}')
        if current_value is not None:
            # Update the database record for the specific room and month
            WaterUsage.query.filter_by(use_year=selected_year, room_no=f'{room_no}').update({
                f'acum_{month_map[selected_month]}': current_value,
            })

    db.session.commit()
    return jsonify({'status': 'success'})

################### Admin 차량 이동거리 리딩 ###################################################
@bp.route('/get_vehicle_data')
def get_vehicle_data():
    year = request.args.get('year')
    month = int(request.args.get('month'))
    prev_year = request.args.get('prev_year')
    prev_month = int(request.args.get('prev_month'))

    month_map = {
        1: "jan",
        2: "feb",
        3: "mar",
        4: "apr",
        5: "may",
        6: "jun",
        7: "jul",
        8: "aug",
        9: "sep",
        10: "oct",
        11: "nov",
        12: "dec"
    }

    month_column = f'acum_{month_map[month]}'  # e.g., acum_jul
    prev_month_column = f'acum_{month_map[prev_month]}'  # e.g., acum_jun

    # 차량 데이터를 담을 딕셔너리
    vehicle_data = {}

    # 차량 번호들을 가져옵니다.
    car_nos = db.session.query(VehicleUsage.car_no).filter(VehicleUsage.use_year == year).distinct().all()

    # 각 차량 번호에 대해 현재 및 이전 달의 데이터를 가져옵니다.
    for car_no_tuple in car_nos:
        car_no = car_no_tuple[0]
        current_record = VehicleUsage.query.filter_by(use_year=year, car_no=car_no).first()
        prev_record = VehicleUsage.query.filter_by(use_year=prev_year, car_no=car_no).first()

        vehicle_data[f'vehicle_current_{car_no}'] = getattr(current_record, month_column, 0) if current_record else 0
        vehicle_data[f'vehicle_prev_{car_no}'] = getattr(prev_record, prev_month_column, 0) if prev_record else 0
    
    print(vehicle_data)
    return jsonify(vehicle_data)



################### Admin 차량 이동거리 입력 #####################################################
@bp.route('/save_vehicle_data', methods=['POST'])
def save_vehicle_data():
    data = request.json
    selected_year = data.get('selected_year')
    selected_month = int(data.get('selected_month'))
    print('데이터는:', data)

    month_map = {
        1: "jan",
        2: "feb",
        3: "mar",
        4: "apr",
        5: "may",
        6: "jun",
        7: "jul",
        8: "aug",
        9: "sep",
        10: "oct",
        11: "nov",
        12: "dec"
    }

    # 차량 번호들을 가져옵니다.
    car_nos = db.session.query(VehicleUsage.car_no).filter(VehicleUsage.use_year == selected_year).distinct().all()

    # 각 차량 번호에 대해 현재 및 이전 달의 데이터를 가져옵니다.
    for car_no_tuple in car_nos:
        car_no = car_no_tuple[0]
        current_value = data.get(f'vehicle_current_{car_no}')
        if current_value is not None:
            # Update the database record for the specific room and month
            VehicleUsage.query.filter_by(use_year=selected_year, car_no=f'{car_no}').update({
                f'acum_{month_map[selected_month]}': current_value,
            })

    db.session.commit()
    return jsonify({'status': 'success'})

################## Admin 차량 등록 #####################################################
@bp.route('/regvehicle', methods=['POST'])
def reg_vehicle():
    data = request.json
    vehicle_no = data.get('vehicle_no')
    vehicle_name = data.get('vehicle_name')
    fuel_type = data.get('fuel_type')
    use_year = data.get('use_year')
    acum_ref = data.get('acum_ref')
    print('데이터는:', data)

    # car_no 중복 확인
    existing_vehicle = VehicleUsage.query.filter_by(car_no=vehicle_no).first()
    if existing_vehicle:
        return jsonify({'status': 'error', 'message': '이미 등록된 차량 번호입니다.'}), 400
    
    # 새로운 차량 정보 등록
    vehicle_usage = VehicleUsage(
        car_no=vehicle_no,
        car_name=vehicle_name,
        car_fuel=fuel_type,
        use_year = use_year,
        acum_ref =acum_ref,
        create_date=datetime.utcnow(),  # 현재 시간으로 생성 날짜 설정
        modify_date=datetime.utcnow()  # 수정 날짜는 처음에는 None으로 설정
    )

    try:
        db.session.add(vehicle_usage)
        db.session.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
        return jsonify({'status': 'error', 'message': '차량 등록 중 오류가 발생했습니다.'}), 500


# ================= data db에 취합 ===============================================
@bp.route('/sumupdate', methods=['GET', 'POST'])
def sum_update():
    # 각 사용량 테이블의 모든 레코드를 가져옵니다.
    water_list = WaterUsage.query.order_by(WaterUsage.no).all()
    elec_list = ElecUsage.query.order_by(ElecUsage.no).all()
    vehicle_list = VehicleUsage.query.order_by(VehicleUsage.no).all()

    # 월별 사용량 합계 계산
    total_use_jan_water = sum([int(item.use_jan) for item in water_list])
    total_use_feb_water = sum([int(item.use_feb) for item in water_list])
    total_use_mar_water = sum([int(item.use_mar) for item in water_list])
    total_use_apr_water = sum([int(item.use_apr) for item in water_list])
    total_use_may_water = sum([int(item.use_may) for item in water_list])
    total_use_jun_water = sum([int(item.use_jun) for item in water_list])
    total_use_jul_water = sum([int(item.use_jul) for item in water_list])
    total_use_aug_water = sum([int(item.use_aug) for item in water_list])
    total_use_sep_water = sum([int(item.use_sep) for item in water_list])
    total_use_oct_water = sum([int(item.use_oct) for item in water_list])
    total_use_nov_water = sum([int(item.use_nov) for item in water_list])
    total_use_dec_water = sum([int(item.use_dec) for item in water_list])

    total_use_jan_elec = sum([int(item.use_jan) for item in elec_list])
    total_use_feb_elec = sum([int(item.use_feb) for item in elec_list])
    total_use_mar_elec = sum([int(item.use_mar) for item in elec_list])
    total_use_apr_elec = sum([int(item.use_apr) for item in elec_list])
    total_use_may_elec = sum([int(item.use_may) for item in elec_list])
    total_use_jun_elec = sum([int(item.use_jun) for item in elec_list])
    total_use_jul_elec = sum([int(item.use_jul) for item in elec_list])
    total_use_aug_elec = sum([int(item.use_aug) for item in elec_list])
    total_use_sep_elec = sum([int(item.use_sep) for item in elec_list])
    total_use_oct_elec = sum([int(item.use_oct) for item in elec_list])
    total_use_nov_elec = sum([int(item.use_nov) for item in elec_list])
    total_use_dec_elec = sum([int(item.use_dec) for item in elec_list])

    total_use_jan_vehicle = sum([int(item.use_jan) for item in vehicle_list])
    total_use_feb_vehicle = sum([int(item.use_feb) for item in vehicle_list])
    total_use_mar_vehicle = sum([int(item.use_mar) for item in vehicle_list])
    total_use_apr_vehicle = sum([int(item.use_apr) for item in vehicle_list])
    total_use_may_vehicle = sum([int(item.use_may) for item in vehicle_list])
    total_use_jun_vehicle = sum([int(item.use_jun) for item in vehicle_list])
    total_use_jul_vehicle = sum([int(item.use_jul) for item in vehicle_list])
    total_use_aug_vehicle = sum([int(item.use_aug) for item in vehicle_list])
    total_use_sep_vehicle = sum([int(item.use_sep) for item in vehicle_list])
    total_use_oct_vehicle = sum([int(item.use_oct) for item in vehicle_list])
    total_use_nov_vehicle = sum([int(item.use_nov) for item in vehicle_list])
    total_use_dec_vehicle = sum([int(item.use_dec) for item in vehicle_list])

    print( total_use_jan_elec)

    # 월별 이름 리스트
    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    # 월별 데이터를 저장하기 위한 루프
    for i, month in enumerate(months):
        # 각 월별 사용량과 CO2 계산
        total_use_elec = locals()[f"total_use_{month}_elec"]
        total_use_water = locals()[f"total_use_{month}_water"]
        total_use_vehicle = locals()[f"total_use_{month}_vehicle"]

        co2_elec = round(total_use_elec * 0.4781, 3)
        co2_water = round(total_use_water * 0.234, 3)
        co2_vehicle = round(total_use_vehicle / 15.35 * 2.582, 3)

        total_co2 = round(co2_elec + co2_water + co2_vehicle, 3)

        # Co2Management 레코드 생성
        new_co2_management = Co2Management(
            use_date=datetime.utcnow().strftime(f'%Y-{i+1:02}'),  # 월을 01, 02, ... 12 형식으로
            use_elec=str(total_use_elec),
            co2_elec=co2_elec,
            use_water=str(total_use_water),
            co2_water=co2_water,
            use_vehicle=str(total_use_vehicle),
            co2_vehicle=co2_vehicle,
            use_total=str(total_co2),
            co2_waste=0.0,  # 아직 사용하지 않는 필드이지만, 0으로 초기화
            use_waste=0.0,  # 아직 사용하지 않는 필드이지만, 0으로 초기화
            co2_gas=0.0,    # 아직 사용하지 않는 필드이지만, 0으로 초기화
            use_gas=0.0,    # 아직 사용하지 않는 필드이지만, 0으로 초기화
            create_date=datetime.utcnow(),
            modify_date=None
        )

        # 데이터베이스에 저장
        db.session.add(new_co2_management)

    # 모든 레코드를 커밋
    db.session.commit()

    return 'ok'


############################## 엑셀 파일 다운로드 업로드 ###########################################3

# 다운로드 FROM DB
@bp.route('/dbdownload', methods = ['GET','POST'])
def excel_download():
    write_wb = Workbook()
    write_ws = write_wb.active
    edata = Stockinfo.query.all()

    ws_title = "종목명 종목코드"
    l_title = ws_title.split("\t")
    write_ws.append(l_title)

    # Append each row of stock info to the worksheet
    for stock_info in edata:
        row_data = (stock_info.stockname, stock_info.stockcode)  # Extract name and code attributes
        write_ws.append(row_data)
    
    file_path = "C:/projects/download/"
    flie_name = "stockcode_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        
    write_wb.save(file_path + flie_name)

    return 'ok'


# 업로드 From excel to DB
@bp.route('/elecupload', methods = ['GET','POST'])
def excel_upload():
    file_path = "C:/projects/co2/rawdata/water_elec/"
    file_name = "water_elec_2024.xlsx"

    wb = openpyxl.load_workbook(file_path + file_name)  # 엑셀 파일을 엽니다.
    ws = wb["elec"]  # "elec" 시트를 선택합니다.

    excel_to_list_all = []  # 엑셀 전체 데이터를 담을 리스트를 초기화합니다.
    seen = set()  # 중복 데이터를 확인하기 위한 집합을 초기화합니다.

    for index, row in enumerate(ws.rows):  # 모든 행을 반복합니다.
        if index >= 2:  # 3번째 행부터 데이터를 저장합니다.
            excel_to_list1 = []  # 한 행의 데이터를 담을 리스트를 초기화합니다.
            duplicate_check = (row[0].value, row[2].value)  # 중복 데이터를 확인하기 위한 튜플을 생성합니다.

            if duplicate_check not in seen:  # 중복이 아닌 경우에만 추가합니다.
                for cell in row:  # 행의 각 셀을 반복합니다.
                    excel_to_list1.append(cell.value)  # 셀의 값을 리스트에 추가합니다.

                excel_to_list_all.append(excel_to_list1)  # 행의 데이터 리스트를 전체 데이터 리스트에 추가합니다.
                seen.add(duplicate_check)  # 중복 데이터 집합에 추가합니다.

    for i in range(len(excel_to_list_all)):  # excel_to_list_all의 길이만큼 반복합니다.
        acum_ref = excel_to_list_all[i][1]
        acum_jan = excel_to_list_all[i][2] if len(excel_to_list_all[i]) > 2 else 0
        acum_feb = excel_to_list_all[i][3] if len(excel_to_list_all[i]) > 3 else 0
        acum_mar = excel_to_list_all[i][4] if len(excel_to_list_all[i]) > 4 else 0
        acum_apr = excel_to_list_all[i][5] if len(excel_to_list_all[i]) > 5 else 0
        acum_may = excel_to_list_all[i][6] if len(excel_to_list_all[i]) > 6 else 0
        acum_jun = excel_to_list_all[i][7] if len(excel_to_list_all[i]) > 7 else 0
        acum_jul = excel_to_list_all[i][8] if len(excel_to_list_all[i]) > 8 else 0
        acum_aug = excel_to_list_all[i][9] if len(excel_to_list_all[i]) > 9 else 0
        acum_sep = excel_to_list_all[i][10] if len(excel_to_list_all[i]) > 10 else 0
        acum_oct = excel_to_list_all[i][11] if len(excel_to_list_all[i]) > 11 else 0
        acum_nov = excel_to_list_all[i][12] if len(excel_to_list_all[i]) > 12 else 0
        acum_dec = excel_to_list_all[i][13] if len(excel_to_list_all[i]) > 13 else 0

         # 필요한 경우 값을 정수로 변환
        acum_ref = int(acum_ref)
        acum_jan = int(acum_jan)
        acum_feb = int(acum_feb)
        acum_mar = int(acum_mar)
        acum_apr = int(acum_apr)
        acum_may = int(acum_may)
        acum_jun = int(acum_jun)
        acum_jul = int(acum_jul)
        acum_aug = int(acum_aug)
        acum_sep = int(acum_sep)
        acum_oct = int(acum_oct)
        acum_nov = int(acum_nov)
        acum_dec = int(acum_dec)

        use_jan = acum_jan - acum_ref if acum_jan > 0 else 0
        use_feb = acum_feb - acum_jan if acum_feb > acum_jan else 0
        use_mar = acum_mar - acum_feb if acum_mar > acum_feb else 0
        use_apr = acum_apr - acum_mar if acum_apr > acum_mar else 0
        use_may = acum_may - acum_apr if acum_may > acum_apr else 0
        use_jun = acum_jun - acum_may if acum_jun > acum_may else 0
        use_jul = acum_jul - acum_jun if acum_jul > acum_jun else 0
        use_aug = acum_aug - acum_jul if acum_aug > acum_jul else 0
        use_sep = acum_sep - acum_aug if acum_sep > acum_aug else 0
        use_oct = acum_oct - acum_sep if acum_oct > acum_sep else 0
        use_nov = acum_nov - acum_oct if acum_nov > acum_oct else 0
        use_dec = acum_dec - acum_nov if acum_dec > acum_nov else 0


        print(use_jan)


        # 총 사용량 계산 (사용된 달들만 합산)
        total = sum([
            use_jan if use_jan >= 0 else 0,
            use_feb if use_feb >= 0 else 0,
            use_mar if use_mar >= 0 else 0,
            use_apr if use_apr >= 0 else 0,
            use_may if use_may >= 0 else 0,
            use_jun if use_jun >= 0 else 0,
            use_jul if use_jul >= 0 else 0,
            use_aug if use_aug >= 0 else 0,
            use_sep if use_sep >= 0 else 0,
            use_oct if use_oct >= 0 else 0,
            use_nov if use_nov >= 0 else 0,
            use_dec if use_dec >= 0 else 0,
        ])
        print(total)

        # 데이터베이스에 저장할 때 유효한 사용량이 있을 때만 추가
        if total > 0:
            elec_usage = ElecUsage(
                room_no=excel_to_list_all[i][0],
                use_year='2024',
                acum_ref=acum_ref,
                acum_jan=acum_jan,
                use_jan=use_jan,
                acum_feb=acum_feb,
                use_feb=use_feb,
                acum_mar=acum_mar,
                use_mar=use_mar,
                acum_apr=acum_apr,
                use_apr=use_apr,
                acum_may=acum_may,
                use_may=use_may,
                acum_jun=acum_jun,
                use_jun=use_jun,
                acum_jul=acum_jul,
                use_jul=use_jul,
                acum_aug=acum_aug,
                use_aug=use_aug,
                acum_sep=acum_sep,
                use_sep=use_sep,
                acum_oct=acum_oct,
                use_oct=use_oct,
                acum_nov=acum_nov,
                use_nov=use_nov,
                acum_dec=acum_dec,
                use_dec=use_dec,
                total=total,
                create_date=datetime.utcnow(),  # 현재 시간으로 생성 날짜 설정
                modify_date=datetime.utcnow()  # 수정 날짜는 처음에는 None으로 설정
            )

            # 데이터베이스에 저장
            db.session.add(elec_usage)
            db.session.commit()

    return 'ok'

# 업로드 From excel to DB
@bp.route('/waterupload', methods = ['GET','POST'])
def water_upload():
    file_path = "C:/projects/co2/rawdata/water_elec/"
    file_name = "water_elec_2024.xlsx"

    wb = openpyxl.load_workbook(file_path + file_name)  # 엑셀 파일을 엽니다.
    ws = wb["water"]  # "elec" 시트를 선택합니다.

    excel_to_list_all = []  # 엑셀 전체 데이터를 담을 리스트를 초기화합니다.
    seen = set()  # 중복 데이터를 확인하기 위한 집합을 초기화합니다.

    for index, row in enumerate(ws.rows):  # 모든 행을 반복합니다.
        if index >= 2:  # 3번째 행부터 데이터를 저장합니다.
            excel_to_list1 = []  # 한 행의 데이터를 담을 리스트를 초기화합니다.
            duplicate_check = (row[0].value, row[2].value)  # 중복 데이터를 확인하기 위한 튜플을 생성합니다.

            if duplicate_check not in seen:  # 중복이 아닌 경우에만 추가합니다.
                for cell in row:  # 행의 각 셀을 반복합니다.
                    excel_to_list1.append(cell.value)  # 셀의 값을 리스트에 추가합니다.

                excel_to_list_all.append(excel_to_list1)  # 행의 데이터 리스트를 전체 데이터 리스트에 추가합니다.
                seen.add(duplicate_check)  # 중복 데이터 집합에 추가합니다.

    for i in range(len(excel_to_list_all)):  # excel_to_list_all의 길이만큼 반복합니다.
        acum_ref = excel_to_list_all[i][1]
        acum_jan = excel_to_list_all[i][2] if len(excel_to_list_all[i]) > 2 else 0
        acum_feb = excel_to_list_all[i][3] if len(excel_to_list_all[i]) > 3 else 0
        acum_mar = excel_to_list_all[i][4] if len(excel_to_list_all[i]) > 4 else 0
        acum_apr = excel_to_list_all[i][5] if len(excel_to_list_all[i]) > 5 else 0
        acum_may = excel_to_list_all[i][6] if len(excel_to_list_all[i]) > 6 else 0
        acum_jun = excel_to_list_all[i][7] if len(excel_to_list_all[i]) > 7 else 0
        acum_jul = excel_to_list_all[i][8] if len(excel_to_list_all[i]) > 8 else 0
        acum_aug = excel_to_list_all[i][9] if len(excel_to_list_all[i]) > 9 else 0
        acum_sep = excel_to_list_all[i][10] if len(excel_to_list_all[i]) > 10 else 0
        acum_oct = excel_to_list_all[i][11] if len(excel_to_list_all[i]) > 11 else 0
        acum_nov = excel_to_list_all[i][12] if len(excel_to_list_all[i]) > 12 else 0
        acum_dec = excel_to_list_all[i][13] if len(excel_to_list_all[i]) > 13 else 0

        use_jan = acum_jan - acum_ref if acum_jan > 0 else 0
        use_feb = acum_feb - acum_jan if acum_feb > acum_jan else 0
        use_mar = acum_mar - acum_feb if acum_mar > acum_feb else 0
        use_apr = acum_apr - acum_mar if acum_apr > acum_mar else 0
        use_may = acum_may - acum_apr if acum_may > acum_apr else 0
        use_jun = acum_jun - acum_may if acum_jun > acum_may else 0
        use_jul = acum_jul - acum_jun if acum_jul > acum_jun else 0
        use_aug = acum_aug - acum_jul if acum_aug > acum_jul else 0
        use_sep = acum_sep - acum_aug if acum_sep > acum_aug else 0
        use_oct = acum_oct - acum_sep if acum_oct > acum_sep else 0
        use_nov = acum_nov - acum_oct if acum_nov > acum_oct else 0
        use_dec = acum_dec - acum_nov if acum_dec > acum_nov else 0


        print(use_jan)


        # 총 사용량 계산 (사용된 달들만 합산)
        total = sum([
            use_jan if use_jan >= 0 else 0,
            use_feb if use_feb >= 0 else 0,
            use_mar if use_mar >= 0 else 0,
            use_apr if use_apr >= 0 else 0,
            use_may if use_may >= 0 else 0,
            use_jun if use_jun >= 0 else 0,
            use_jul if use_jul >= 0 else 0,
            use_aug if use_aug >= 0 else 0,
            use_sep if use_sep >= 0 else 0,
            use_oct if use_oct >= 0 else 0,
            use_nov if use_nov >= 0 else 0,
            use_dec if use_dec >= 0 else 0,
        ])
        print(total)

        # 데이터베이스에 저장할 때 유효한 사용량이 있을 때만 추가
        if total > 0:
            water_usage = WaterUsage(
                room_no=excel_to_list_all[i][0],
                use_year='2024',
                acum_ref=acum_ref,
                acum_jan=acum_jan,
                use_jan=use_jan,
                acum_feb=acum_feb,
                use_feb=use_feb,
                acum_mar=acum_mar,
                use_mar=use_mar,
                acum_apr=acum_apr,
                use_apr=use_apr,
                acum_may=acum_may,
                use_may=use_may,
                acum_jun=acum_jun,
                use_jun=use_jun,
                acum_jul=acum_jul,
                use_jul=use_jul,
                acum_aug=acum_aug,
                use_aug=use_aug,
                acum_sep=acum_sep,
                use_sep=use_sep,
                acum_oct=acum_oct,
                use_oct=use_oct,
                acum_nov=acum_nov,
                use_nov=use_nov,
                acum_dec=acum_dec,
                use_dec=use_dec,
                total=total,
                create_date=datetime.utcnow(),  # 현재 시간으로 생성 날짜 설정
                modify_date=datetime.utcnow()  # 수정 날짜는 처음에는 None으로 설정
            )

            # 데이터베이스에 저장
            db.session.add(water_usage)
            db.session.commit()

    return 'ok'

# 업로드 From excel to DB
@bp.route('/vehicleupload', methods = ['GET','POST'])
def vehicle_upload():
    file_path = "C:/projects/co2/rawdata/vehicle/"
    file_name = "vehicle.xlsx"

    wb = openpyxl.load_workbook(file_path + file_name)  # 엑셀 파일을 엽니다.
    ws = wb["vehicle"]  # "elec" 시트를 선택합니다.

    excel_to_list_all = []  # 엑셀 전체 데이터를 담을 리스트를 초기화합니다.
    seen = set()  # 중복 데이터를 확인하기 위한 집합을 초기화합니다.

    for index, row in enumerate(ws.rows):  # 모든 행을 반복합니다.
        if index >= 2:  # 3번째 행부터 데이터를 저장합니다.
            excel_to_list1 = []  # 한 행의 데이터를 담을 리스트를 초기화합니다.
            duplicate_check = (1, 4)  # 중복 데이터를 확인하기 위한 튜플을 생성합니다.

            if duplicate_check not in seen:  # 중복이 아닌 경우에만 추가합니다.
                for cell in row:  # 행의 각 셀을 반복합니다.
                    excel_to_list1.append(cell.value)  # 셀의 값을 리스트에 추가합니다.

                excel_to_list_all.append(excel_to_list1)  # 행의 데이터 리스트를 전체 데이터 리스트에 추가합니다.
                seen.add(duplicate_check)  # 중복 데이터 집합에 추가합니다.

    for i in range(len(excel_to_list_all)):  # excel_to_list_all의 길이만큼 반복합니다.
        acum_ref = excel_to_list_all[i][3]
        acum_jan = excel_to_list_all[i][4] if len(excel_to_list_all[i]) > 2 else 0
        acum_feb = excel_to_list_all[i][5] if len(excel_to_list_all[i]) > 3 else 0
        acum_mar = excel_to_list_all[i][6] if len(excel_to_list_all[i]) > 4 else 0
        acum_apr = excel_to_list_all[i][7] if len(excel_to_list_all[i]) > 5 else 0
        acum_may = excel_to_list_all[i][8] if len(excel_to_list_all[i]) > 6 else 0
        acum_jun = excel_to_list_all[i][9] if len(excel_to_list_all[i]) > 7 else 0
        acum_jul = excel_to_list_all[i][10] if len(excel_to_list_all[i]) > 8 else 0
        acum_aug = excel_to_list_all[i][11] if len(excel_to_list_all[i]) > 9 else 0
        acum_sep = excel_to_list_all[i][12] if len(excel_to_list_all[i]) > 10 else 0
        acum_oct = excel_to_list_all[i][13] if len(excel_to_list_all[i]) > 11 else 0
        acum_nov = excel_to_list_all[i][14] if len(excel_to_list_all[i]) > 12 else 0
        acum_dec = excel_to_list_all[i][15] if len(excel_to_list_all[i]) > 13 else 0


        # 필요한 경우 값을 정수로 변환
        acum_ref = int(acum_ref)
        acum_jan = int(acum_jan)
        acum_feb = int(acum_feb)
        acum_mar = int(acum_mar)
        acum_apr = int(acum_apr)
        acum_may = int(acum_may)
        acum_jun = int(acum_jun)
        acum_jul = int(acum_jul)
        acum_aug = int(acum_aug)
        acum_sep = int(acum_sep)
        acum_oct = int(acum_oct)
        acum_nov = int(acum_nov)
        acum_dec = int(acum_dec)

        use_jan = acum_jan - acum_ref if acum_jan > 0 else 0
        use_feb = acum_feb - acum_jan if acum_feb > acum_jan else 0
        use_mar = acum_mar - acum_feb if acum_mar > acum_feb else 0
        use_apr = acum_apr - acum_mar if acum_apr > acum_mar else 0
        use_may = acum_may - acum_apr if acum_may > acum_apr else 0
        use_jun = acum_jun - acum_may if acum_jun > acum_may else 0
        use_jul = acum_jul - acum_jun if acum_jul > acum_jun else 0
        use_aug = acum_aug - acum_jul if acum_aug > acum_jul else 0
        use_sep = acum_sep - acum_aug if acum_sep > acum_aug else 0
        use_oct = acum_oct - acum_sep if acum_oct > acum_sep else 0
        use_nov = acum_nov - acum_oct if acum_nov > acum_oct else 0
        use_dec = acum_dec - acum_nov if acum_dec > acum_nov else 0


        print(use_jan)


        # 총 사용량 계산 (사용된 달들만 합산)
        total = sum([
            use_jan if use_jan >= 0 else 0,
            use_feb if use_feb >= 0 else 0,
            use_mar if use_mar >= 0 else 0,
            use_apr if use_apr >= 0 else 0,
            use_may if use_may >= 0 else 0,
            use_jun if use_jun >= 0 else 0,
            use_jul if use_jul >= 0 else 0,
            use_aug if use_aug >= 0 else 0,
            use_sep if use_sep >= 0 else 0,
            use_oct if use_oct >= 0 else 0,
            use_nov if use_nov >= 0 else 0,
            use_dec if use_dec >= 0 else 0,
        ])
        print(total)

        # 데이터베이스에 저장할 때 유효한 사용량이 있을 때만 추가
        if total > 0:
            vehicle_usage = VehicleUsage(
                car_no=excel_to_list_all[i][0],
                car_name=excel_to_list_all[i][1],
                car_fuel=excel_to_list_all[i][2],
                use_year='2024',
                acum_ref=acum_ref,
                acum_jan=acum_jan,
                use_jan=use_jan,
                acum_feb=acum_feb,
                use_feb=use_feb,
                acum_mar=acum_mar,
                use_mar=use_mar,
                acum_apr=acum_apr,
                use_apr=use_apr,
                acum_may=acum_may,
                use_may=use_may,
                acum_jun=acum_jun,
                use_jun=use_jun,
                acum_jul=acum_jul,
                use_jul=use_jul,
                acum_aug=acum_aug,
                use_aug=use_aug,
                acum_sep=acum_sep,
                use_sep=use_sep,
                acum_oct=acum_oct,
                use_oct=use_oct,
                acum_nov=acum_nov,
                use_nov=use_nov,
                acum_dec=acum_dec,
                use_dec=use_dec,
                total=total,
                create_date=datetime.utcnow(),  # 현재 시간으로 생성 날짜 설정
                modify_date=datetime.utcnow()  # 수정 날짜는 처음에는 None으로 설정
            )

            # 데이터베이스에 저장
            db.session.add(vehicle_usage)
            db.session.commit()

    return 'ok'

# #===================== SortExcel daily VB Data  20240514~~ ==================================
# # 업로드 From excel to DB
# @bp.route('/sortexcel', methods = ['GET','POST'])
# def excel_sort():

#     minus_time = request.args.get('time', type=int, default=2)

#     # 오늘,어제, 현재시간(6자리),현재시간(hhmm00 분봉용) 호출 6자리()
#     todate,ydate,totime,totime00 = KFunction.date_info(minus_time)  

#     file_path = "//DESKTOP-F5S9HG9/공유 with PC/주마등/검증프로그램/보관/"
#     file_name = "Stock_Find Max(rev5.18)_" + ydate +".xlsm"

#     print(file_name)     

#     # Read the Excel file starting from the 7th row and read data from the third sheet
#     #df = pd.read_excel(file_path + file_name, header=None, skiprows=6, names=['Column1', 'Column2', 'Column3'], sheet_name=2) # 모드칼럼
#     #칼럼 1~3번까지만
#     df = pd.read_excel(file_path + file_name, usecols=[0, 1, 2], header=None, skiprows=6, names=['Column1', 'Column2', 'Column3'], sheet_name=2)

#     #print(df)
#     # 중복된 항목은 20240516 기준으로 하나만 남김
#     #filtered_df = df[(df['Column3'] == 20240516) & (df['Column1'].duplicated() == False)]

#     print(df)

#     # Find the maximum date value in Column3
#     max_date = df['Column3'].max()

#     # Filter the DataFrame to include only values with the maximum date value
#     filtered_df = df[~df['Column1'].isin(df[df['Column3'] < max_date]['Column1'])]

#     # Further exclude duplicates based on the values in Column1
#     filtered_df = filtered_df.drop_duplicates(subset='Column1')
#     print(filtered_df)
#     for index, row in filtered_df.iterrows():
#         daily_data_fromvb = DailyDataFromVB(
#             # stockcode 값에서 앞에 'A'가 있는 경우 제거하고 사용
#             stockcode = row['Column1'].lstrip('A'),
#             stockname=row['Column2'],
#             stockdate=row['Column3'],
#             fromvb = '1'
#         )

#         # 데이터베이스에 저장
#         db.session.add(daily_data_fromvb)
#         db.session.commit()
 
#     return redirect(url_for('kinvestor._dailylist'))
