from datetime import datetime, timedelta
from openpyxl import Workbook
import openpyxl    
import asyncio
import time
import pandas as pd
import matplotlib.pyplot as plt
import base64
import mplfinance as mpf
#from pylab import *
import os
import io
#from aioflask import Flask, request, jsonify, Blueprint, render_template,url_for
from flask import Flask, request, jsonify, Blueprint, render_template,url_for, send_file
from werkzeug.utils import redirect
from ..models import Stockinfo, Chartinfo, StockinfoTemp,Token_temp, ObserveStock, StockListFromVB, ObserveStockFromVB, DailyDataFromVB, StockListFromKInvestor,ThemeCode,StockBalanceLists
import json
from .. import db
from sqlalchemy import func, desc, and_ , Table, MetaData, Column, Integer, String, select, update, insert

# from flask_sqlalchemy import pagination
from sqlalchemy.orm import aliased
from ..service.k_trade import KTrade
from ..service.k_analysis import KAnalysis
from ..service.k_value import KValue
from ..service.k_realdata import RealTimeData,RealData
from ..service.k_db_trade import KDatabaseTrade
from ..service.k_func import KFunction
from ..service.k_table import KTempTable
from collections import Counter 


import websockets
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from base64 import b64decode

# bp = Flask(__name__)

bp = Blueprint('kinvestor',__name__,url_prefix='/kinvestor')

# 'Agg' 백엔드 사용
plt.switch_backend('Agg')

# 현재 실행 중인 연결 객체
current_connection = None

#plt 한글설정
plt.rcParams['font.family'] = 'Malgun Gothic'



# approvalkey와 token 얻기
@bp.route('/connection', methods=['GET', 'POST', 'DELETE'])
def token_key():
    key_lists = []

    #urls for approval key & token (real, mock(simulated))
    urlkey = 'https://openapi.koreainvestment.com:9443/oauth2/Approval'
    urlkey_mock = 'https://openapivts.koreainvestment.com:29443/oauth2/Approval'

    urltoken = 'https://openapi.koreainvestment.com:9443/oauth2/tokenP'
    urltoken_mock = 'https://openapivts.koreainvestment.com:29443/oauth2/tokenP'

    with open("C:/projects/daram/kinvestor_keys.txt") as f:
        app_keys = f.read()
        #print(app_keys)
        key_lists.append(app_keys[app_keys.find('app_key="')+9:app_keys.find('"=app_key')])
        key_lists.append(app_keys[app_keys.find('app_secret="')+12:app_keys.find('"=app_secret')])
        key_lists.append(app_keys[app_keys.find('app_key_mock="')+14:app_keys.find('"=app_key_mock')])
        key_lists.append(app_keys[app_keys.find('app_secret_mock="')+17:app_keys.find('"=app_secret_mock')])
    
    #print(key_lists)  
        
    # 승인키 획득
    approval_key = KTrade.approvalKey(key_lists[0],key_lists[1],urlkey)
    approval_key_mock =  KTrade.approvalKey(key_lists[2],key_lists[3],urlkey_mock)
    
    if approval_key and approval_key_mock:
        #토큰 획득
        token, token_expire= KTrade.tokenP(key_lists[0],key_lists[1],urltoken)
        token_mock, token_expire_mock= KTrade.tokenP(key_lists[2],key_lists[3],urltoken_mock)
        if token and token_mock:
            #기존 DB값 삭제
            Token_temp.query.delete()
            db.session.commit()
            # Chartinfo 객체 생성 및 속성 채우기
            tokenDB = Token_temp(
                approval_key = approval_key,
                token_key = token,
                token_expire = token_expire,
                approval_key_mock = approval_key_mock,
                token_key_mock = token_mock,
                token_expire_mock = token_expire_mock     
            )
            # 데이터베이스에 저장
            db.session.add(tokenDB)
            db.session.commit()
           
            return redirect(url_for('kinvestor._list'))
        else:
           return "token 획득 실패, 1분 후 다시 시도해주시요"
    else:
        return "approval_key 획득 실패, 인터넷 상태 점검해주세요"

# token 호출 (실계좌)
def call_token():
     # 토근 DB로 부터 호출
    token = Token_temp.query.first()
   
    # 주어진 시간 문자열을 datetime 객체로 변환
    given_time = datetime.strptime(token.token_expire, "%Y-%m-%d %H:%M:%S")
    # 현재 시간
    current_time = datetime.now()

    # 현재 시간과 주어진 시간 사이의 차이 계산
    time_difference =  given_time - current_time 
    print(time_difference, given_time, current_time)

    # 현재 시간보다 2시간 전인지 비교
    two_hours_ago = timedelta(hours=2)

    if time_difference > two_hours_ago:
        static_token = token.token_key
        return static_token
    else:
        #redirect(url_for('kinvestor.token_key'))
        token_key()

# token 호출 (모의 계좌)
def call_token_mock():
     # 토근 DB로 부터 호출
    token = Token_temp.query.first()
   
    # 주어진 시간 문자열을 datetime 객체로 변환
    given_time = datetime.strptime(token.token_expire_mock, "%Y-%m-%d %H:%M:%S")
    # 현재 시간
    current_time = datetime.now()

    # 현재 시간과 주어진 시간 사이의 차이 계산
    time_difference =  given_time - current_time 
    print(time_difference, given_time, current_time)

    # 현재 시간보다 2시간 전인지 비교
    two_hours_ago = timedelta(hours=2)

    if time_difference > two_hours_ago:
        static_token = token.token_key_mock
        return static_token
    else:
        token_key()

# 종목 코드 조회   
@bp.route('/stocklist/')
def _list():
    #입력파라메터
    page = request.args.get('page', type=int, default=1) #페이지
    kw = request.args.get('kw',type=str, default='')
    so = request.args.get('so', type=str, default='recent')

    # #조회
    # 토근 DB로 부터 호출
    # token = Token_temp.query.first()
    # static_token = token.token_key
    static_token = call_token_mock()

    stock_list =  Stockinfo.query.order_by(Stockinfo.stockcode) 
    json_data = KValue.current_value("000150",static_token)
   # print(json_data)
    # 조회
    
    if kw:
        search = '%%{}%%'.format(kw)
        stock_list = db.session.query(Stockinfo).filter(
        (Stockinfo.stockcode.ilike(search)) | # 주식 코드
        (Stockinfo.stockname.ilike(search)) # 주식 이름
        ).distinct()
 # 페이징
    stock_list = stock_list.paginate(page=page, per_page=10)
    return render_template('stockdata/stock_list.html', stock_list=stock_list, page=page, kw=kw)



#주식 일별 차트 정보 자세히
@bp.route('/detail', methods=['GET'])
def stock_detail():
    ki = KValue()

    # 입력 파라미터
    page = request.args.get('page', type=int, default=1)
    kw = request.args.get('kw', type=str, default='')
    so = request.args.get('so', type=str, default='recent')

    stockcode1 = request.args.get('stockcode')
    stockcode = stockcode1.replace('A','')
    n = request.args.get('n', type=str)
    date_from = request.args.get('from', type=str)
    date_to = request.args.get('to', type=str)
    
    dailydatas = [] # output2 일별데이터
    stockinfo = [] #output1 종목 정보
    # 토큰 가져오기
    static_token = call_token_mock()

    # 파라미터 처리
    if n == '':
        n = None
    if date_from == '' and date_to == '':
        date_from = None
        date_to = None

    if date_from is not None and date_to is not None:
        date_from = date_from.replace('-', '')
        date_to = date_to.replace('-', '')

    if n is None and date_from is None and date_to is None:
        # 현재 날짜를 'YYYYMMDD' 형식의 문자열로 가져오기
        date_to = datetime.today().strftime('%Y%m%d')

        # date_to를 기준으로 120일 전의 날짜 계산
        date_to_dt = datetime.strptime(date_to, '%Y%m%d')
        date_from_dt = date_to_dt - timedelta(days=120)

        # date_from을 'YYYYMMDD' 형식의 문자열로 변환
        date_from = date_from_dt.strftime('%Y%m%d')
        print(f"date_from: {date_from}, date_to: {date_to}")

        # 초기 요청
        stockinfo,stockcandles = ki.period_price(stockcode, static_token, date_from, date_to)
        # stockcandles_dic = json.loads(scandles)
        # stockcandles = stockcandles_dic["output2"]
        # stockinfo = stockcandles_dic["output1"]

        # dailydatas 초기화
        dailydatas = stockcandles

        # 데이터가 정상적으로 수집되었으면 추가 요청
        if stockcandles:
            # data_to - 1일을 구함
            date_to_dt = date_from_dt - timedelta(days=1)
            date_to = date_to_dt.strftime('%Y%m%d')

            # data_from - 120일을 구함
            date_from_dt = date_to_dt - timedelta(days=120)
            date_from = date_from_dt.strftime('%Y%m%d')
            print(f"date_from: {date_from}, date_to: {date_to}")

            # 다시 한 번 데이터 요청
            stockinfo,stockcandles = ki.period_price(stockcode, static_token, date_from, date_to)
            # stockcandles_dic = json.loads(scandles)
            # stockcandles = stockcandles_dic["output2"]
            # stockinfo = stockcandles_dic["output1"]

            print('스탁인포:', stockinfo)

            # 추가 데이터를 리스트에 합병
            dailydatas.extend(stockcandles)

        print(dailydatas)

    if not (n or date_from):
        return 'Need to provide "n" or "date_from" argument.', 400

    # # 토큰 가져오기
    # static_token = call_token_mock()

    # # 주식 캔들 데이터 가져오기
    # scandles = ki.period_price(stockcode, static_token, date_from, date_to)
    

    # 데이터베이스 처리
    StockinfoTemp.query.delete()
    db.session.commit()

    # Chartinfo 객체 생성 및 속성 채우기
    stockinfor = StockinfoTemp(
            stockcode= stockinfo['stck_shrn_iscd'],
            stockname=stockinfo['hts_kor_isnm'],
            currentvalue =stockinfo['stck_prpr'],
            highvalue  =stockinfo['stck_hgpr'] ,
            lowvalue =stockinfo['stck_lwpr'],
            beginvalue =stockinfo['stck_oprc'],
            diffrate  =stockinfo['prdy_ctrt'], #등락률
            diffval =stockinfo['prdy_vrss'], #전일대비 
            tradeval =stockinfo['acml_vol'], #현재 거래량
            pre_tradeval =stockinfo['prdy_vol'], #전일거래량
            tvol_vsprevious =stockinfo['prdy_vrss_vol'], # 전일대비 거래량
            faceval =stockinfo['stck_fcam'], #액면가
            stockvol =stockinfo['lstn_stcn'], #상장주수
            capital =stockinfo['cpfn'], #자본금
            stocksum =stockinfo['hts_avls'], #시가총액
            per =stockinfo['per'], #per
            eps =stockinfo['eps'] ,# eps
            pbr =stockinfo['pbr'], #pbr
            debtrate =stockinfo['itewhol_loan_rmnd_ratem name'] # 전체 융자 잔고 비율
        )

    db.session.add(stockinfor)
    db.session.commit()

    Chartinfo.query.delete()
    db.session.commit()

    for data in dailydatas:
        date = data['stck_bsop_date']
        beginval = str(int(data['stck_oprc']))
        highval = str(int(data['stck_hgpr']))
        lowval = str(int(data['stck_lwpr']))
        endval = str(int(data['stck_clpr']))
        tradeval = str(data['acml_vol'])
        tradevol = str(data['acml_tr_pbmn'])
        changeval = str(int(data['prdy_vrss']))
        csign = data['prdy_vrss_sign']

        if csign == '2':
            changesign = "+"
        elif csign == '5':
            changesign = "-"
        else:
            changesign = "0"

        changevol = str(round(int(changeval) / (int(endval) - int(changeval)) * 100, 1)) + "%"

        chartinfo = Chartinfo(
            stockcode=stockcode,
            date=date,
            beginval=beginval,
            highval=highval,
            lowval=lowval,
            endval=endval,
            changeval=changeval,
            tradeval=tradeval,
            tradevol=tradevol,
            changesign=changesign,
            changevol=changevol,
        )

        db.session.add(chartinfo)
        db.session.commit()

    chart_list = Chartinfo.query.order_by(Chartinfo.date.desc())
    stock_info = StockinfoTemp.query.all()


    # 분봉을 위한 테이블 이름 구성
    temp_min_table_name = 'table_min_' + stockcode
    metadata = MetaData()
    table = Table(temp_min_table_name, metadata, autoload_with=db.engine)
    #저장된 분봉 데이터 요청
    min_datas = db.session.query(table).order_by(table.c.tradetime.desc())
  
    gragh_list = []
    for chart_info in chart_list:
        row = [
            chart_info.date,
            int(chart_info.beginval),
            int(chart_info.highval),
            int(chart_info.lowval),
            int(chart_info.endval),
            int(chart_info.tradeval),
        ]
        gragh_list.append(row)
    print("일데이터:",gragh_list)
    gragh_min_list = []
    for chart_info in min_datas:
        row = [
            chart_info.tradetime,
            int(chart_info.open),
            int(chart_info.high),
            int(chart_info.low),
            int(chart_info.close),
            int(chart_info.tradevol),
        ]
        gragh_min_list.append(row)
    print("민데이터:",gragh_min_list)

    #print(gragh_list)

    #image_url = url_for('kinvestor.get_chart_image', _external=True)
    image_url = get_chart_image(gragh_list,'day')
    image_min_url = get_chart_image(gragh_min_list,'min')
    #print('이미지url',image_url)



    chart_count = Chartinfo.query.count()
    avg_5, avg_20, avg_60, avg_120, avg_200 = None, None, None, None, None

    if chart_count >= 5:
        avg_5 = db.session.query(func.avg(Chartinfo.endval)).filter(Chartinfo.no.between(1, 5)).scalar()
    if chart_count >= 20:
        avg_20 = db.session.query(func.avg(Chartinfo.endval)).filter(Chartinfo.no.between(1, 20)).scalar()
    if chart_count >= 60:
        avg_60 = db.session.query(func.avg(Chartinfo.endval)).filter(Chartinfo.no.between(1, 60)).scalar()
    if chart_count >= 120:
        avg_120 = db.session.query(func.avg(Chartinfo.endval)).filter(Chartinfo.no.between(1, 120)).scalar()
    if chart_count >= 200:
        avg_200 = db.session.query(func.avg(Chartinfo.endval)).filter(Chartinfo.no.between(1, 200)).scalar()

    if kw:
        search = '%%{}%%'.format(kw)
        stock_list = db.session.query(Chartinfo).filter(
            (Chartinfo.stockcode.ilike(search)) |
            (Chartinfo.stockname.ilike(search))
        ).distinct()

    chart_list = chart_list.paginate(page=page, per_page=20)

    return render_template('stockdata/stock_detail.html', 
                            image_url=image_url,
                            image_min_url = image_min_url,
                            chart_list=chart_list, 
                            stock_info=stock_info[0] if stock_info else None,
                            page=page, 
                            kw=kw, 
                            total_pages=chart_count, 
                            stockcode=stockcode)

# 차트 그래프###################################################3
# 차트 그래프 생성 함수
def get_chart_image(gragh_list, method):
    import pandas as pd
    import matplotlib.pyplot as plt
    import mplfinance as mpf
    import io
    import base64
    
    # 리스트를 데이터프레임으로 변환
    columns = ['Date', 'Open', 'High', 'Low', 'Close', 'Volume']
    data_df = pd.DataFrame(gragh_list, columns=columns)
   
    # 날짜 형식을 변환하고 인덱스 설정
    if method == 'day':
        data_df['Date'] = pd.to_datetime(data_df['Date'], format='%Y%m%d')
        data_df.set_index('Date', inplace=True)
        data_df = data_df[::-1]  # 시간순으로 데이터 정렬
        print("일데이터:", data_df)
    elif method == 'min':
        # Example assuming 'Date' is in format 'HHMM00' as string
        data_df['Date'] = pd.to_datetime('19000101' + data_df['Date'], format='%Y%m%d%H%M%S')
        data_df.set_index('Date', inplace=True)
        data_df = data_df[::-1]  # 시간순으로 데이터 정렬
        print("민데이터: ", data_df)

    # 데이터 리샘플링 함수 정의
    def resample_data(data, period):
        if period == 'D':
            return data
        elif period == 'W':
            return data.resample('W').agg({'Open': 'first', 'High': 'max', 'Low': 'min', 'Close': 'last', 'Volume': 'sum'}).dropna()
        elif period == 'M':
            return data.resample('M').agg({'Open': 'first', 'High': 'max', 'Low': 'min', 'Close': 'last', 'Volume': 'sum'}).dropna()

    period = 'D'
    resampled_data = resample_data(data_df, period)

    # 이동평균선 설정
    moving_averages = {
        '3_MA': 3,
        '5_MA': 5,
        '10_MA': 10,
        '20_MA': 20,
        '60_MA': 60,
        '120_MA': 120
    }

    # 각 이동평균선 계산
    for ma_name, window in moving_averages.items():
        resampled_data[ma_name] = resampled_data['Close'].rolling(window=window).mean()

    # 최대 데이터 표시 개수 설정 (최대 120개)
    max_display = min(len(resampled_data), 120)
    display_data = resampled_data.tail(max_display)

    print("디스플레이데이터:", display_data)

    # NaN 값을 이전 값으로 채우기 (forward fill)
    display_data.fillna(method='ffill', inplace=True)

    # NaN 값이 포함된 경우 해당 MA는 제외하고 추가할 plot 설정
    apds = []
    colors = {'3_MA': 'red', '5_MA': 'blue', '10_MA': 'green', '20_MA': 'orange', '60_MA': 'purple', '120_MA': 'brown'}
    for ma_name, window in moving_averages.items():
        if ma_name in display_data.columns and display_data[ma_name].notna().any():
            apds.append(mpf.make_addplot(display_data[ma_name], color=colors[ma_name], linestyle='-', width=0.7, label=ma_name))
    
    # 한글 폰트 설정
    plt.rcParams['font.family'] = 'Malgun Gothic'  # 한글 폰트 설정
    mc = mpf.make_marketcolors(up='r', down='b', inherit=True)
    s = mpf.make_mpf_style(marketcolors=mc, gridstyle='dotted')

    # 그래프 그리기
    fig, axlist = mpf.plot(display_data, type='candle', style=s, addplot=apds, volume=True, 
                           title='Daily Stock Chart  (3_MA: red, 5_MA: blue, 10_MA: Green, 20_MA: Orange, 60_MA: violet, 120_MA: brown)', ylabel='Price', ylabel_lower='Volume', 
                           show_nontrading=False, returnfig=True, figsize=(15, 7))  # 그래프 크기 설정 (인치 단위)

    # Add legend below the title
    axlist[0].legend(loc='upper center', bbox_to_anchor=(0.5, -0.05), ncol=6, fontsize='small')

    # Add annotations
    annotations = '3_MA: red, 5_MA: blue, 10_MA: Green, 20_MA: Orange, 60_MA: violet, 120_MA: brown'
    plt.annotate(annotations, xy=(0.5, 1.05), xycoords='axes fraction', ha='center', fontsize='small', color='black')

    # 그래프를 바이트 버퍼로 저장
    buf = io.BytesIO()
    fig.savefig(buf, format='png')
    buf.seek(0)

    # 그래프 메모리 해제
    plt.close(fig)

    # 이미지를 URL로 변환 (사용자 정의에 따라 수정 필요)
    image_url = convert_image_to_url(buf.getvalue())

    return image_url




        # Return the image file as a response (직접 client에게 보낼때)
        #return send_file(buf, mimetype='image/png')

def convert_image_to_url(image_data):
    encoded_img = base64.b64encode(image_data).decode('utf-8')
    image_url = f'data:image/png;base64,{encoded_img}'
    return image_url




#========================================================================================#
# 관심종목 20
#=========================================================================================#
@bp.route('/observe', methods = ['GET','POST'])
async def observe_stock():
    
    kt = KTrade()
    kv = KValue()
    kd =RealData()

        #입력파라메터
    page = request.args.get('page', type=int, default=1) #페이지
    kw = request.args.get('kw',type=str, default='')
    so = request.args.get('so', type=str, default='recent')
    trade_type = request.args.get('trade', type=str, default='mock')
    if request.method == 'POST':
        print('POST')
    else:
        o_stock = ObserveStock.query.all()
        static_token = call_token_mock()
        current_value = []
        code_list = []
        for value in o_stock:
            stockcode = value.stockcode.replace('A','') # 한투는 숫자 6자리만 인식함으로 앞자리 삭제

            #trade_type = 'mock'  # mock : 모의투자 , real : 실투자
            #실호가,실체결가,체결통보
            #code_list = [['1','H0STASP0',stockcode],['1','H0STCNT0',stockcode],['1','H0STCNI9','njsk2002']] # 모의: H0STCNI9 실:H0STCNI0
            c_list = ['1','H0STCNT0',stockcode] # 모의: H0STCNI9 실:H0STCNI0
            code_list.append(c_list)
            #현재가는 속도가 늦어서 사용 보류
            #c_data = KValue.current_value(stockcode,static_token)
        print(code_list)
        #rdata = await kd.connect(code_list,trade_type)
  
        rdata = asyncio.create_task(kd.connect(code_list, trade_type))
        await asyncio.sleep(30)
        kd.stop()
        print(rdata)
        rdata = asyncio.create_task(kd.connect(code_list, trade_type))
        await asyncio.sleep(30)
        kd.stop()
        # curr_data = 'json.loads(c_data)'
        # #print('현재가 데이터', curr_data)
        # current_data = curr_data["output"]
            
        # stock_info = {
        #             "stockcode": value.stockcode,
        #             "stockname": value.stockname,
        #             "cvalue": current_data['stck_prpr'],
        #             "svalue": current_data['stck_oprc'],
        #             "hvalue": current_data['stck_hgpr'],
        #             "lvalue": current_data['stck_lwpr'],
        #             "tradeval": current_data['acml_vol'],
        #             "diffval": current_data['prdy_vrss'],
        #             "diffrate": current_data['prdy_ctrt']
        #         }

        # current_value.append(stock_info)
        current_value = '10'
        time.sleep(100/1000)
       # print(current_value)      
        return render_template('stockdata/observe_list.html', current_value=current_value)

#========================================================================================#
# 관심종목 20  from vb
#=========================================================================================#
@bp.route('/observefromvb', methods=['GET', 'POST'])
async def observe_fromvb():

    kt = KTrade()
    kv = KValue()
    kr = RealTimeData()
    kd = RealData()

    # 입력파라메터
    page = request.args.get('page', type=int, default=1)  # 페이지
    kw = request.args.get('kw', type=str, default='')
    so = request.args.get('so', type=str, default='recent')
    r_method = request.args.get('r_method', type=str, default='continue')
    if request.method == 'POST':
        print('POST')
    else:
       
        # 현재 날짜와 시간을 datetime 객체로 얻어옵니다.
        today_date = datetime.now()

        # datetime 객체를 원하는 형식의 문자열로 변환합니다.
        today_date_str = today_date.strftime("%Y%m%d")

        # 문자열을 datetime 객체로 다시 변환합니다. 필요한 경우에만 사용합니다.
        today_date_dt = datetime.strptime(today_date_str, "%Y%m%d")

        # 현재 날짜에서 하루를 빼어 어제의 날짜를 구합니다.
        
        yesterday_date = today_date_dt - timedelta(days=1)



        # 어제의 날짜를 원하는 형식의 문자열로 변환합니다.
        ydate = yesterday_date.strftime("%Y%m%d")
        print(ydate)
        trade_type = 'mock'  # mock : 모의투자 , real : 실투자
        # SQLAlchemy ORM을 사용하여 데이터베이스에서 데이터를 가져옵니다.
        #query = StockListFromVB.query.filter_by(stockdate=ydate)
        stockcodes = DailyDataFromVB.query.filter_by(selected1 = '1')

        #######################################################################3
        ## 처음에만 StockListFromVB에서 전일날자기준 데이터를 한번만 가져오고,
        ## 이후부터는 코드에 채결데이터(또는 1분차트를 저장해서) 계속 업데이터 시킴.
        

        #ObserveStockFromVB 이전데이터 삭제후 fromvb의 전일데이터만 업데이트
        # ObserveStockFromVB.query.delete()
        # db.session.commit()
        # for que in query:
        #     ob_fromvb = ObserveStockFromVB(
        #         stockcode = que.stockcode.replace('A', ''),
        #         stockname= que.stockname,
        #         stockdate= que.stockdate,
        #         method_1 = que.method_1,
        #         method_2 = que.method_2
        #     )
        #     db.session.add(ob_fromvb)
        #     db.session.commit()

        # query = ObserveStockFromVB.query.filter_by(stockdate=ydate)

        ObserveStockFromVB.query.delete()
        db.session.commit()
        for que in stockcodes:
            ob_fromvb = ObserveStockFromVB(
                stockcode = que.stockcode.replace('A', ''),
                stockname= que.stockname,
                stockdate= que.stockdate,
                method_1 = que.selected1,
                method_2 = today_date_dt  #매수 날짜
            )
            db.session.add(ob_fromvb)
            db.session.commit()

        query = ObserveStockFromVB.query.filter_by(method_1='1')


        # 페이징 처리
        query = query.paginate(page=page, per_page=10)
        print(query.items)
        static_token = call_token_mock()
        current_value = []
        code_list = []
        for value in query.items:
            stockcode = value.stockcode.replace('A', '')  # 한투는 숫자 6자리만 인식함으로 앞자리 삭제
            print("스탁코드는: ", stockcode)

            # 실호가,실체결가,체결통보
            # code_list = [['1','H0STASP0',stockcode],['1','H0STCNT0',stockcode],['1','H0STCNI9','njsk2002']] # 모의: H0STCNI9 실:H0STCNI0
            c_list = ['1', 'H0STCNT0', stockcode]  # 모의: H0STCNI9 실:H0STCNI0
            code_list.append(c_list)
            # 현재가는 속도가 늦어서 사용 보류
            # c_data = KValue.current_value(stockcode,static_token)
        print(code_list)

        # 페이지가 1보다 큰 경우에는 새로운 종목 코드 리스트를 가져옴
        if page > 1:
            # 현재 페이지에서 가져올 종목 코드 리스트를 생성
            new_code_list = []
            for value in query.items:
                stockcode = value.stockcode.replace('A', '')  # 한투는 숫자 6자리만 인식함으로 앞자리 삭제
                c_list = ['1', 'H0STCNT0', stockcode]  # 모의: H0STCNI9 실:H0STCNI0
                new_code_list.append(c_list)

            # 이전에 실행 중인 connect 메서드를 중지
            await kd.stop()
            #asyncio.kd.get_event_loop().close()
            print("제발스탑 스탑 스탑")

            # 새로운 종목 코드 리스트로 connect 메서드 실행
            #rdata = await asyncio.create_task(kd.connect(new_code_list, trade_type, r_method))
            #print("알데이터는 = ", rdata)
        else:
            # 페이지가 1인 경우에는 이전에 사용한 코드 리스트로 connect 메서드 실행
            #rdata = await asyncio.create_task(kd.connect(code_list, trade_type, r_method))
            await kd.start(code_list, trade_type, r_method)
        current_value = '10'
        time.sleep(100 / 1000)

        if kw:
            search = '%%{}%%'.format(kw)
            stock_list = db.session.query(Stockinfo).filter(
                (Stockinfo.stockcode.ilike(search)) |  # 주식 코드
                (Stockinfo.stockname.ilike(search))  # 주식 이름
            ).distinct()
        # 페이징

        #return render_template('stockdata/observe_list.html', stock_list=rdata, page=page, kw=kw)
        return 'ok'


#========================================================================================#
# Dailydata를 Observe로 이동 
#=========================================================================================#
@bp.route('/observefromdaily', methods=['GET', 'POST'])
def observe_fromdaily():

    kt = KTrade()
    kv = KValue()
    kr = RealTimeData()
    kd = RealData()

    # 입력파라메터
    page = request.args.get('page', type=int, default=1)  # 페이지
    kw = request.args.get('kw', type=str, default='')
    so = request.args.get('so', type=str, default='recent')
    r_method = request.args.get('r_method', type=str, default='continue')
    if request.method == 'POST':
        print('POST')
    else:
       
        # 오늘,어제, 현재시간(6자리),현재시간(hhmm00 분봉용) 호출 6자리()
        todate,ydate,totime,totime00 = KFunction.date_info(2)  

        # SQLAlchemy ORM을 사용하여 데이터베이스에서 데이터를 가져옵니다.
        #query = StockListFromVB.query.filter_by(stockdate=ydate)
        # 첫 번째 방식: 두 번의 필터링을 체인
        # stockcodes = DailyDataFromVB.query.filter_by(selected1='1').filter(DailyDataFromVB.selecteddate == ydate).all()
        # print(stockcodes)  # 결과 출력

        # 두 번째 방식: filter_by 메소드에서 여러 조건을 한 번에 적용
        stockcodes = DailyDataFromVB.query.filter_by(selected1='1', selecteddate=ydate).all()  # .all()을 호출해야 쿼리가 실행됩니다
        print(stockcodes)  # 결과 출력

        #기존 데이블 삭제
        KTempTable.drop_min_table('all','observe')

        ObserveStockFromVB.query.delete()
        db.session.commit()
        for que in stockcodes:
            ob_fromvb = ObserveStockFromVB(
                stockcode = que.stockcode.replace('A', ''),
                stockname= que.stockname,
                stockdate= que.stockdate,
                method_1 = que.selected1,
                selected1 = que.selected1,
                selecteddate = que.selecteddate,
                method_2 = todate, #매수 날짜
                buydate = totime  #매수 시간
            )
            db.session.add(ob_fromvb)
            db.session.commit()
            #임시테이블 추가
            KTempTable.create_min_table(que.stockcode.replace('A', ''),'observe')


        query = ObserveStockFromVB.query.filter_by(method_1='1')

    return redirect(url_for('kinvestor.observe_fromdb'))
        

# AJAX 비동기 ######  daily데이터에서 Sorting된 데이터 
@bp.route('/observefromdb2', methods=['GET', 'POST'])
def observe_fromdb2():
    if request.method == 'POST':
        print('POST')
        # page = request.args.get('page', type=int, default=1)
        # kw = request.args.get('kw', type=str, default='')
        # so = request.args.get('so', type=str, default='recent')
        page = request.form.get('page', type=int, default=1)
        kw = request.form.get('kw', type=str, default='')
        so = request.form.get('so', type=str, default='recent')
        print(page)
        static_token = call_token_mock()
        if kw:
            search = '%%{}%%'.format(kw)
            stock_list = db.session.query(ObserveStockFromVB).filter(
                (ObserveStockFromVB.stockcode.ilike(search)) | # 주식 코드
                (ObserveStockFromVB.stockname.ilike(search))   # 주식 이름
            ).distinct().paginate(page=page, per_page=10).items # 페이징
        
        # 검색어(kw)가 없는 경우 모든 데이터를 가져옵니다.
        else:
            stock_codes = ObserveStockFromVB.query.with_entities(ObserveStockFromVB.stockcode).all()
            
            c_data = []
            for stock_code in stock_codes:
                # stock_code는 튜플 형태 (stockcode,) 이므로 첫 번째 요소만 가져와야 합니다.
                stock_code = stock_code[0]
                print(stock_code)
                # KValue의 current_value 함수를 사용하여 각 stock_code에 대한 현재 값을 가져옵니다.
                time.sleep(0.3)
                current_value = KValue.current_value(stock_code, static_token)

                print("current_value: ", current_value)

                # 주어진 stock_code에 대한 항목이 이미 있는지 확인합니다
                observe_fromvb = db.session.query(ObserveStockFromVB).filter_by(stockcode=stock_code).first()

                if observe_fromvb:
                    # 항목이 존재하면 해당 필드를 업데이트합니다
                    observe_fromvb.currentvalue = current_value['stck_prpr']
                    observe_fromvb.beginvalue = current_value['stck_oprc']
                    observe_fromvb.highvalue = current_value['stck_hgpr']
                    observe_fromvb.lowvalue = current_value['stck_lwpr']
                    observe_fromvb.tradeval = current_value['acml_vol']
                    observe_fromvb.tradevalrate = current_value['prdy_vrss_vol_rate']
                    observe_fromvb.diffval = current_value['prdy_vrss']
                    observe_fromvb.diffrate = current_value['prdy_ctrt']
                    observe_fromvb.caution = current_value['invt_caful_yn']
                    observe_fromvb.warning = current_value['mrkt_warn_cls_code']
                else:
                    # 항목이 존재하지 않으면 새 항목을 생성합니다
                    observe_fromvb = ObserveStockFromVB(
                        stock_code=stock_code,  # stock_code를 저장하는 것을 잊지 마세요
                        currentvalue=current_value['stck_prpr'],
                        beginvalue=current_value['stck_oprc'],
                        highvalue=current_value['stck_hgpr'],
                        lowvalue=current_value['stck_lwpr'],
                        tradeval=current_value['acml_vol'],  # 누적거래량
                        tradevalrate=current_value['prdy_vrss_vol_rate'],  # 전일 거래량 대비
                        diffval=current_value['prdy_vrss'],  # 전일종가대비
                        diffrate=current_value['prdy_ctrt'],  # 전일종가대비비율
                        caution=current_value['invt_caful_yn'],  # 투자유의여부
                        warning=current_value['mrkt_warn_cls_code']  # 00: 없음 01: 투자주의 02: 투자경고 03: 투자위험
                    )
                    db.session.add(observe_fromvb)

                # 세션을 커밋하여 변경 사항을 저장합니다
                db.session.commit()


            
            #stock_list = ObserveStockFromVB.query.paginate(page=page, per_page=20).items # pagination data를 제외한 순수데이터만 얻을때
            stock_list = ObserveStockFromVB.query.paginate(page=page, per_page=20)

       

        # return jsonify(
        #     page=stock_list.page,
        #     pagination={
        #         'page': stock_list.page,
        #         'per_page': stock_list.per_page,
        #         'total_pages': stock_list.pages,
        #         'total_items': stock_list.total,
        #         'has_next': stock_list.has_next,
        #         'has_prev': stock_list.has_prev,
        #         'next_num': stock_list.next_num,
        #         'prev_num': stock_list.prev_num
        #     },
        #     stock_list=result
        # )

        # JSON으로 데이터를 반환합니다.
        return jsonify(
            page=page,
            pagination={
                'page': stock_list.page,
                'per_page': stock_list.per_page,
                'total_pages': stock_list.pages,
                'total_items': stock_list.total,
                'has_next': stock_list.has_next,
                'has_prev': stock_list.has_prev,
                'next_num': stock_list.next_num,
                'prev_num': stock_list.prev_num
            },
            stock_list=[{
                'stocknum' :stock.no,
                'stockcode': stock.stockcode,
                'stockname': stock.stockname,
                'stockdate': stock.stockdate,
                'currentvalue': stock.currentvalue,
                'beginvalue': stock.beginvalue,
                'highvalue': stock.highvalue,
                'lowvalue': stock.lowvalue,
                'tradeval': stock.tradeval,
                'diffval': stock.diffval,
                'diffrate': stock.diffrate
            } for stock in stock_list.items]
        )

    else:
        page = request.args.get('page', type=int, default=1)
        kw = request.args.get('kw', type=str, default='')
        so = request.args.get('so', type=str, default='recent')
    
        query = ObserveStockFromVB.query
        query = query.paginate(page=page, per_page=20)

        print(query)

        return render_template('stockdata/observe_list.html', stock_list=query, page=page, kw=kw)
    

# AJAX 비동기 ######  daily데이터에서 Sorting된 데이터 / ### 분봉데이터 ++++ ####
@bp.route('/observefromdb', methods=['GET', 'POST'])
def observe_fromdb():
    if request.method == 'POST':
        print('POST')
        page = request.form.get('page', type=int, default=1)
        kw = request.form.get('kw', type=str, default='')
        so = request.form.get('so', type=str, default='recent')
        print(page)
        static_token_mock = call_token_mock()
        static_token = call_token()

        # 오늘,어제, 현재시간(6자리), 현재시간(hhmm00 분봉용) 호출 6자리()
        todate, ydate, totime, totime00 = KFunction.date_info(0)  

        if kw:
            search = '%%{}%%'.format(kw)
            stock_list = db.session.query(ObserveStockFromVB).filter(
                (ObserveStockFromVB.stockcode.ilike(search)) |  # 주식 코드
                (ObserveStockFromVB.stockname.ilike(search))   # 주식 이름
            ).distinct().paginate(page=page, per_page=10).items  # 페이징
        
        else:
            stock_codes = ObserveStockFromVB.query.with_entities(ObserveStockFromVB.stockcode).all()
            
            c_data = []
            if totime00 >= '090000':
                for s_code in stock_codes:
                    stock_code = s_code[0]
                    print(stock_code)
                    
                    todate, ydate, totime, totime00 = KFunction.date_info(0)  
                    
                    # 테이블 이름 구성
                    temp_min_table_name = 'table_min_' + stock_code
                    metadata = MetaData()
                    table = Table(temp_min_table_name, metadata, autoload_with=db.engine)
                    
                    data_min_list = []
                    if totime00 < '093000':
                        data_min = KValue.mindata(stock_code, static_token_mock, totime00)
                        print("totime001:", totime00)
                        data_min_list.extend(data_min)
                    elif totime00 >= '093000' and totime00 <= '235000':
                        totime, totime00, before_times = KFunction.beforetime_info(30)
                        min_data = db.session.query(table.c.tradetime).filter_by(tradetime=before_times[0]).all()
                        print("before_time: ", before_times[0])
                        if min_data:
                            data_min = KValue.mindata(stock_code, static_token_mock, totime00)
                            data_min_list.extend(data_min)
                        else:
                            totime, totime00, before_times = KFunction.beforetime_info(0)
                            print("시간데이터:", before_times)
                            for beforetime00 in before_times:
                                data_min = KValue.mindata(stock_code, static_token_mock, beforetime00)
                                data_min_list.extend(data_min)
                    
                    print("min데이터:", data_min_list)

                    for min_data in data_min_list:
                        record = db.session.query(table).filter_by(tradetime=min_data['stck_cntg_hour']).first()
                        print('시간: ',min_data)
                        
                        if record:
                            print('recode.open:',min_data['stck_cntg_hour'],record.open,min_data['stck_oprc'])
                            update_statement = (
                                update(table).
                                where(table.c.tradetime == min_data['stck_cntg_hour']).
                                values(
                                    open=min_data['stck_oprc'],
                                    high=min_data['stck_hgpr'],
                                    low=min_data['stck_lwpr'],
                                    close=min_data['stck_prpr'],
                                    tradevol=min_data['cntg_vol']
                                )
                            )
                            db.session.execute(update_statement)
                        else:
                            insert_statement = table.insert().values(
                                tradetime=min_data['stck_cntg_hour'],
                                open=min_data['stck_oprc'],
                                high=min_data['stck_hgpr'],
                                low=min_data['stck_lwpr'],
                                close=min_data['stck_prpr'],
                                tradevol=min_data['cntg_vol']
                            )
                            db.session.execute(insert_statement)

                        db.session.commit()
                        
                    current_value = KValue.current_value(stock_code, static_token_mock)

                    #print("current_value: ", current_value)

                    # 주어진 stock_code에 대한 항목이 이미 있는지 확인합니다
                    observe_fromvb = db.session.query(ObserveStockFromVB).filter_by(stockcode=stock_code).first()

                    if observe_fromvb:
                        # 항목이 존재하면 해당 필드를 업데이트합니다
                        observe_fromvb.currentvalue = current_value['stck_prpr']
                        observe_fromvb.beginvalue = current_value['stck_oprc']
                        observe_fromvb.highvalue = current_value['stck_hgpr']
                        observe_fromvb.lowvalue = current_value['stck_lwpr']
                        observe_fromvb.tradeval = current_value['acml_vol']
                        observe_fromvb.tradevalrate = current_value['prdy_vrss_vol_rate']
                        observe_fromvb.diffval = current_value['prdy_vrss']
                        observe_fromvb.diffrate = current_value['prdy_ctrt']
                        observe_fromvb.caution = current_value['invt_caful_yn']
                        observe_fromvb.warning = current_value['mrkt_warn_cls_code']
                    else:
                        # 항목이 존재하지 않으면 새 항목을 생성합니다
                        observe_fromvb = ObserveStockFromVB(
                            stock_code=stock_code,  # stock_code를 저장하는 것을 잊지 마세요
                            currentvalue=current_value['stck_prpr'],
                            beginvalue=current_value['stck_oprc'],
                            highvalue=current_value['stck_hgpr'],
                            lowvalue=current_value['stck_lwpr'],
                            tradeval=current_value['acml_vol'],  # 누적거래량
                            tradevalrate=current_value['prdy_vrss_vol_rate'],  # 전일 거래량 대비
                            diffval=current_value['prdy_vrss'],  # 전일종가대비
                            diffrate=current_value['prdy_ctrt'],  # 전일종가대비비율
                            caution=current_value['invt_caful_yn'],  # 투자유의여부
                            warning=current_value['mrkt_warn_cls_code']  # 00: 없음 01: 투자주의 02: 투자경고 03: 투자위험
                        )
                        db.session.add(observe_fromvb)

                    # 세션을 커밋하여 변경 사항을 저장합니다
                    db.session.commit()


            
            #stock_list = ObserveStockFromVB.query.paginate(page=page, per_page=20).items # pagination data를 제외한 순수데이터만 얻을때
            stock_list = ObserveStockFromVB.query.paginate(page=page, per_page=20)

       

        # return jsonify(
        #     page=stock_list.page,
        #     pagination={
        #         'page': stock_list.page,
        #         'per_page': stock_list.per_page,
        #         'total_pages': stock_list.pages,
        #         'total_items': stock_list.total,
        #         'has_next': stock_list.has_next,
        #         'has_prev': stock_list.has_prev,
        #         'next_num': stock_list.next_num,
        #         'prev_num': stock_list.prev_num
        #     },
        #     stock_list=result
        # )

        # JSON으로 데이터를 반환합니다.
        return jsonify(
            page=page,
            pagination={
                'page': stock_list.page,
                'per_page': stock_list.per_page,
                'total_pages': stock_list.pages,
                'total_items': stock_list.total,
                'has_next': stock_list.has_next,
                'has_prev': stock_list.has_prev,
                'next_num': stock_list.next_num,
                'prev_num': stock_list.prev_num
            },
            stock_list=[{
                'stocknum' :stock.no,
                'stockcode': stock.stockcode,
                'stockname': stock.stockname,
                'stockdate': stock.stockdate,
                'currentvalue': stock.currentvalue,
                'beginvalue': stock.beginvalue,
                'highvalue': stock.highvalue,
                'lowvalue': stock.lowvalue,
                'tradeval': stock.tradeval,
                'diffval': stock.diffval,
                'diffrate': stock.diffrate
            } for stock in stock_list.items]
        )

    else:
        page = request.args.get('page', type=int, default=1)
        kw = request.args.get('kw', type=str, default='')
        so = request.args.get('so', type=str, default='recent')
    
        query = ObserveStockFromVB.query
        query = query.paginate(page=page, per_page=20)

        print(query)

        return render_template('stockdata/observe_list.html', stock_list=query, page=page, kw=kw)
    
#####################################################################################################
# AJAX 비동기 ###### 
# Stock_Remainder로 현재 보유중인 주식 잔고 상태 확인 가능
#####################################################################################################
@bp.route('/stockremainders', methods=['GET', 'POST'])
def stock_remainders():
    if request.method == 'POST':
        print('POST')
        # page = request.args.get('page', type=int, default=1)
        # kw = request.args.get('kw', type=str, default='')
        # so = request.args.get('so', type=str, default='recent')
        page = request.form.get('page', type=int, default=1)
        kw = request.form.get('kw', type=str, default='')
        so = request.form.get('so', type=str, default='recent')
        print(page)
        static_token = call_token_mock()

        return_val = 'all' # 모든항목 리턴
        if kw:
            search = '%%{}%%'.format(kw)
            stock_list = db.session.query(StockBalanceLists).filter(
                (StockBalanceLists.stockcode.ilike(search)) | # 주식 코드
                (StockBalanceLists.stockname.ilike(search))   # 주식 이름
            ).distinct().paginate(page=page, per_page=10).items # 페이징
        
        # 검색어(kw)가 없는 경우 모든 데이터를 가져옵니다.
        else:
            # 주식 잔고 db에 업데이트
            stock_codes = KDatabaseTrade.remainder_lists(static_token,"",return_val)  #return_val = 'all'임에 따라, stock_code 필요 없음.
            #stock_codes = StockBalanceLists.query.with_entities(StockBalanceLists.stockcode).all()
            #print(stock_codes)
            c_data = []
            for stock_code in stock_codes:
                # stock_code는 튜플 형태 (stockcode,) 이므로 첫 번째 요소만 가져와야 합니다.
                
                stock_code = stock_code.stockcode
                print(stock_code)
                # KValue의 current_value 함수를 사용하여 각 stock_code에 대한 현재 값을 가져옵니다.
                time.sleep(0.5)
                current_value = KValue.current_value(stock_code, static_token)

                print("current_value: ", current_value)

                # 주어진 stock_code에 대한 항목이 이미 있는지 확인합니다
                stock_balance_list = db.session.query(StockBalanceLists).filter_by(stockcode=stock_code).first()

                if stock_balance_list:
                    # 항목이 존재하면 해당 필드를 업데이트합니다
                    stock_balance_list.currvalue = current_value['stck_prpr']
                    stock_balance_list.beginvalue = current_value['stck_oprc']
                    stock_balance_list.highvalue = current_value['stck_hgpr']
                    stock_balance_list.lowvalue = current_value['stck_lwpr']
                    stock_balance_list.tradeval = current_value['acml_vol']
                    stock_balance_list.tradevalrate = current_value['prdy_vrss_vol_rate']
                    stock_balance_list.diffval = current_value['prdy_vrss']
                    stock_balance_list.diffrate = current_value['prdy_ctrt']
                    stock_balance_list.caution = current_value['invt_caful_yn']
                    stock_balance_list.warning = current_value['mrkt_warn_cls_code']
                else:
                    # 항목이 존재하지 않으면 새 항목을 생성합니다
                    stock_balance_list = StockBalanceLists(
                        stock_code=stock_code,  # stock_code 꼭 생성
                        currentvalue=current_value['stck_prpr'],
                        beginvalue=current_value['stck_oprc'],
                        highvalue=current_value['stck_hgpr'],
                        lowvalue=current_value['stck_lwpr'],
                        tradeval=current_value['acml_vol'],  # 누적거래량
                        tradevalrate=current_value['prdy_vrss_vol_rate'],  # 전일 거래량 대비
                        diffval=current_value['prdy_vrss'],  # 전일종가대비
                        diffrate=current_value['prdy_ctrt'],  # 전일종가대비비율
                        caution=current_value['invt_caful_yn'],  # 투자유의여부
                        warning=current_value['mrkt_warn_cls_code']  # 00: 없음 01: 투자주의 02: 투자경고 03: 투자위험
                    )
                    db.session.add(stock_balance_list)

                # 세션을 커밋하여 변경 사항을 저장합니다
                db.session.commit()

            # 주식의 evalpriceamount 총합 계산
            total_evalpriceamount = db.session.query(func.sum(StockBalanceLists.evalpriceamount)).scalar()
            
            #stock_list = StockBalanceLists.query.paginate(page=page, per_page=20).items # pagination data를 제외한 순수데이터만 얻을때
           
            stock_list = StockBalanceLists.query.paginate(page=page, per_page=15)

       

        # return jsonify(
        #     page=stock_list.page,
        #     pagination={
        #         'page': stock_list.page,
        #         'per_page': stock_list.per_page,
        #         'total_pages': stock_list.pages,
        #         'total_items': stock_list.total,
        #         'has_next': stock_list.has_next,
        #         'has_prev': stock_list.has_prev,
        #         'next_num': stock_list.next_num,
        #         'prev_num': stock_list.prev_num
        #     },
        #     stock_list=result
        # )

        # JSON으로 데이터를 반환합니다.
        return jsonify(
            page=page,
            pagination={
                'page': stock_list.page,
                'per_page': stock_list.per_page,
                'total_pages': stock_list.pages,
                'total_items': stock_list.total,
                'has_next': stock_list.has_next,
                'has_prev': stock_list.has_prev,
                'next_num': stock_list.next_num,
                'prev_num': stock_list.prev_num
            },
            total_evalpriceamount=total_evalpriceamount,  # evalpriceamount 총합 추가
            stock_list=[{
                'stocknum' :stock.no,
                'stockcode': stock.stockcode,
                'stockname': stock.stockname,
                'remainderqty': stock.remainderqty,
                'buyprice': stock.buyprice,
                'buyamount': stock.buyamount,
                'evalrate': stock.evalrate,
                'evalamount': stock.evalpriceamount,
                'currentvalue': stock.currvalue,
                'beginvalue': stock.beginvalue,
                'highvalue': stock.highvalue,
                'lowvalue': stock.lowvalue,
                'tradeval': stock.tradeval,
                'diffval': stock.diffval,
                'diffrate': stock.diffrate
            } for stock in stock_list.items]
        )

    else:
        page = request.args.get('page', type=int, default=1)
        kw = request.args.get('kw', type=str, default='')
        so = request.args.get('so', type=str, default='recent')
    
        query = StockBalanceLists.query
        query = query.paginate(page=page, per_page=10)

        print(query)

    return render_template('stockdata/stock_remainders.html', stock_list=query, page=page, kw=kw)



############ 분봉 요청 ####################################################
@bp.route('/minlist', methods=['GET', 'POST'])
def minlist():
    #입력파라메터
    page = request.args.get('page', type=int, default=1) #페이지
    kw = request.args.get('kw',type=str, default='')
    so = request.args.get('so', type=str, default='recent')
    if request.method == 'POST':
        print('POST')
    else:
       
        # 현재 날짜와 시간을 datetime 객체로 얻어옵니다.
        today_date = datetime.now()

        # datetime 객체를 원하는 형식의 문자열로 변환합니다.
        today_date_str = today_date.strftime("%Y%m%d")

        # 문자열을 datetime 객체로 다시 변환합니다. 필요한 경우에만 사용합니다.
        today_date_dt = datetime.strptime(today_date_str, "%Y%m%d")

        # 현재 날짜에서 하루를 빼어 어제의 날짜를 구합니다.
        yesterday_date = today_date_dt - timedelta(days=1)

        # 어제의 날짜를 원하는 형식의 문자열로 변환합니다.
        ydate = yesterday_date.strftime("%Y%m%d")
        print(ydate)
        trade_type = 'mock'  # mock : 모의투자 , real : 실투자
        # SQLAlchemy ORM을 사용하여 데이터베이스에서 데이터를 가져옵니다.
        query = StockListFromVB.query.filter_by(stockdate=ydate)
        #######################################################################3
        ## 처음에만 StockListFromVB에서 전일날자기준 데이터를 한번만 가져오고,
        ## 이후부터는 코드에 채결데이터(또는 1분차트를 저장해서) 계속 업데이터 시킴.
        

        #ObserveStockFromVB 이전데이터 삭제후 fromvb의 전일데이터만 업데이트
        ObserveStockFromVB.query.delete()
        db.session.commit()
        for que in query:
            ob_fromvb = ObserveStockFromVB(
                stockcode = que.stockcode.replace('A', ''),
                stockname= que.stockname,
                stockdate= que.stockdate,
                method_1 = que.method_1,
                method_2 = que.method_2
            )
            db.session.add(ob_fromvb)
            db.session.commit()

        query = ObserveStockFromVB.query.filter_by(stockdate=ydate)


        # 페이징 처리
        query = query.paginate(page=page, per_page=40)
        print(query.items)
        static_token = call_token_mock()
        current_value = []
        code_list = []
        for value in query.items:
            stockcode = value.stockcode.replace('A', '')  # 한투는 숫자 6자리만 인식함으로 앞자리 삭제
            print("스탁코드는: ", stockcode)

            # 실호가,실체결가,체결통보
            # code_list = [['1','H0STASP0',stockcode],['1','H0STCNT0',stockcode],['1','H0STCNI9','njsk2002']] # 모의: H0STCNI9 실:H0STCNI0
            c_list = ['1', 'H0STCNT0', stockcode]  # 모의: H0STCNI9 실:H0STCNI0
            code_list.append(c_list)
            # 현재가는 속도가 늦어서 사용 보류
            # c_data = KValue.current_value(stockcode,static_token)
        print(code_list)

    # #조회
    # 토근 DB로 부터 호출
    # token = Token_temp.query.first()
    # static_token = token.token_key
    static_token = call_token_mock()

    query = ObserveStockFromVB.query.filter_by(stockdate=ydate)
    #json_data = KValue.current_value("000150",static_token)
   # print(json_data)
    # 조회
    
    if kw:
        search = '%%{}%%'.format(kw)
        stock_list = db.session.query(Stockinfo).filter(
        (Stockinfo.stockcode.ilike(search)) | # 주식 코드
        (Stockinfo.stockname.ilike(search)) # 주식 이름
        ).distinct()
 # 페이징
    stock_list = stock_list.paginate(page=page, per_page=10)
    return render_template('stockdata/stock_list.html', stock_list=stock_list, page=page, kw=kw)


# @bp.route('/observefromvb', methods=['GET', 'POST'])
# async def observe_fromvb():
#     global current_connection

#     if request.method == 'POST':
#         print('POST')
#     else:
#         page = request.args.get('page', type=int, default=1)
#         kw = request.args.get('kw', type=str, default='')
#         so = request.args.get('so', type=str, default='recent')

#         # 이전 연결 중지
#         if current_connection:
#             await current_connection.aclose()  # 현재 async generator를 종료합니다.
#             current_connection = None

#         current_connection = await create_connection(page, kw, so)

#         return render_template('stockdata/observe_list.html', data=current_connection)


# # 페이지에 따라 다른 연결을 생성하는 함수
# async def create_connection(page, kw, so):
#     kt = KTrade()
#     kv = KValue()
#     kd = RealData()

#     today_date = datetime.now()
#     today_date_str = today_date.strftime("%Y%m%d")
#     today_date_dt = datetime.strptime(today_date_str, "%Y%m%d")
#     yesterday_date = today_date_dt - timedelta(days=1)
#     ydate = yesterday_date.strftime("%Y%m%d")

#     trade_type = 'mock'
#     query = StockListFromVB.query.filter_by(stockdate=ydate)
#     ObserveStockFromVB.query.delete()
#     db.session.commit()
#     for que in query:
#         ob_fromvb = ObserveStockFromVB(
#             stockcode = que.stockcode.replace('A', ''),
#             stockname= que.stockname,
#             stockdate= que.stockdate,
#             method_1 = que.method_1,
#             method_2 = que.method_2
#         )
#         db.session.add(ob_fromvb)
#         db.session.commit()

#     query = query.paginate(page=page, per_page=10)
#     static_token = call_token_mock()
#     current_value = []
#     code_list = []

#     for value in query.items:
#         stockcode = value.stockcode.replace('A', '')
#         c_list = ['1', 'H0STCNT0', stockcode]
#         code_list.append(c_list)

#     # 비동기로 connect() 메서드 호출
#     rdata = await kd.connect(code_list, trade_type)

#     current_value = '10'
#     time.sleep(100/1000)

#     if kw:
#         search = '%%{}%%'.format(kw)
#         stock_list = db.session.query(Stockinfo).filter(
#             (Stockinfo.stockcode.ilike(search)) |
#             (Stockinfo.stockname.ilike(search))
#         ).distinct()

#     # async generator를 반환합니다.
#     async def generate_data():
#         for data in rdata:
#             yield data

#     return generate_data()


    



#========================================================================================#
# 관심종목 20 DB에 값 전송
@bp.route('/inputobserve', methods = ['GET','POST'])
def input_observe():

    # ka = KAnalysis
    # static_token = call_token_mock()
    # high_rate = ka.high_rate(static_token)
    
    #임시로 스탁리스트 사용
    stock_list =  Stockinfo.query.order_by(Stockinfo.stockcode) 

    for i, highdata in enumerate(stock_list):
        if (i <20):
            stockcode = highdata.stockcode
            stockname = highdata.stockname

            stockinfo = ObserveStock(stockcode=stockcode,stockname=stockname,create_date=datetime.now(),modify_date=datetime.now())
            db.session.add(stockinfo)
            db.session.commit()
    return redirect(url_for('kinvestor.observe_stock'))
#========================================================================================#
# 주식 매수 / 매도 
#=========================================================================================#
@bp.route('/traderun', methods = ['GET','POST'])
def trade_run():
    kt = KTrade
    

    if request.method == 'POST':
        stockcode = request.args.get('stockcode')
        buyqty = request.args.get('buyqty', type=str)  
        buyprice= request.args.get('buyprice', type=str)
        omethod = request.args.get('ordermethod', type=str)
        #토근 호출 
        static_token = call_token_mock()
        
        if omethod == '00':
            ordermethod = "VTTC0802U"
        elif omethod =='11':
            ordermethod = "VTTC0801U"
        #매수/매도 함수 호출
        #result = kt. order(stockcode,buyqty,buyprice, static_token,ordermethod)
    static_token = call_token_mock()
    result = kt. order("068270","30","187000", static_token,"VTTC0801U")

        #[실전투자]
                #TTTC0802U : 주식 현금 매수 주문
                #TTTC0801U : 주식 현금 매도 주문
                #[모의투자]
                #VTTC0802U : 주식 현금 매수 주문
                #VTTC0801U : 주식 현금 매도 주문
    print(result)
        
    return 'ok'


############################## 엑셀 파일 다운로드 업로드 ###########################################3

# 다운로드 FROM DB
@bp.route('/dbdownload', methods = ['GET','POST'])
def excel_download():
    write_wb = Workbook()
    write_ws = write_wb.active
    edata = Stockinfo.query.all()

    ws_title = "종목명 종목코드"
    l_title = ws_title.split("\t")
    write_ws.append(l_title)

    # Append each row of stock info to the worksheet
    for stock_info in edata:
        row_data = (stock_info.stockname, stock_info.stockcode)  # Extract name and code attributes
        write_ws.append(row_data)
    
    file_path = "C:/projects/download/"
    flie_name = "stockcode_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        
    write_wb.save(file_path + flie_name)

    return 'ok'


# 업로드 From excel to DB
@bp.route('/dbupload', methods = ['GET','POST'])
def excel_upload():
    file_path = "C:/projects/download/"
    file_name = "Stock_Find Max(rev5.18)_20240516.xlsm"

    wb = openpyxl.load_workbook(file_path + file_name)  # 엑셀 파일을 엽니다.
    ws = wb["Sheet3"]  # "Sheet3" 시트를 선택합니다.

    excel_to_list_all = []  # 엑셀 전체 데이터를 담을 리스트를 초기화합니다.

    seen = set()  # 중복 데이터를 확인하기 위한 집합을 초기화합니다.

    for index, row in enumerate(ws.rows):  # 모든 행을 반복합니다.
        if index >= 6:  # 7번째 행부터 데이터를 저장합니다.
            excel_to_list1 = []  # 한 행의 데이터를 담을 리스트를 초기화합니다.
            duplicate_check = (row[0].value, row[2].value)  # 중복 데이터를 확인하기 위한 튜플을 생성합니다.

            if duplicate_check not in seen:  # 중복이 아닌 경우에만 추가합니다.
                for cell in row:  # 행의 각 셀을 반복합니다.
                    excel_to_list1.append(cell.value)  # 셀의 값을 리스트에 추가합니다.

                excel_to_list_all.append(excel_to_list1)  # 행의 데이터 리스트를 전체 데이터 리스트에 추가합니다.
                seen.add(duplicate_check)  # 중복 데이터 집합에 추가합니다.

    print(excel_to_list_all[0][1])  # 전체 데이터 리스트를 출력합니다.

    # Chartinfo 객체 생성 및 속성 채우기

    for i in range(len(excel_to_list_all) - 1):  # excel_to_list_all의 길이만큼 반복합니다.
        stocklist_fromvb = StockListFromVB(
            stockcode= excel_to_list_all[i][0],
            stockname= excel_to_list_all[i][1],
            stockdate= excel_to_list_all[i][2],
            method_1 = excel_to_list_all[i][3],
            method_2 = excel_to_list_all[i][4],
            currentvalue = excel_to_list_all[i][6],
            d5d20  = excel_to_list_all[i][7],
            close5d =excel_to_list_all[i][8],
            close20d = excel_to_list_all[i][9],
            closebegin  = excel_to_list_all[i][10],
            closepreclose = excel_to_list_all[i][11],
            closelow = excel_to_list_all[i][12],
            highclose = excel_to_list_all[i][13],
            begin_1 = excel_to_list_all[i][23],
            begin_2 = excel_to_list_all[i][24],
            begin_3 = excel_to_list_all[i][25],
            begin_4 = excel_to_list_all[i][26],
            begin_5 = excel_to_list_all[i][27],
            high_1 = excel_to_list_all[i][28],
            high_2 = excel_to_list_all[i][29],
            high_3 = excel_to_list_all[i][30],
            high_4 = excel_to_list_all[i][31],
            high_5 = excel_to_list_all[i][32],
            first_low_date = excel_to_list_all[i][33],
            first_low_value = excel_to_list_all[i][34],
            first_high_date = excel_to_list_all[i][35],
            first_high_value = excel_to_list_all[i][36],
            secend_low_date = excel_to_list_all[i][37],
            secend_low_value = excel_to_list_all[i][38],
            secend_high_date = excel_to_list_all[i][39],
            secend_high_value = excel_to_list_all[i][40],
            high_begin_4 = excel_to_list_all[i][41],
            high_begin_3 = excel_to_list_all[i][42],
            high_begin_2 = excel_to_list_all[i][43],
            high_begin_1 = excel_to_list_all[i][44],
            tradevol_4= excel_to_list_all[i][45],
            tradevol_3 = excel_to_list_all[i][46],
            tradevol_2 = excel_to_list_all[i][47],
            tradevol_1 = excel_to_list_all[i][48],
            tradevol_0 = excel_to_list_all[i][49],
            d60_d20 = excel_to_list_all[i][50],
            d20_d5 = excel_to_list_all[i][51],
            close_close4 = excel_to_list_all[i][52]
               
            )

        # 데이터베이스에 저장
        db.session.add(stocklist_fromvb)
        db.session.commit()

    return 'ok'

#===================== SortExcel daily VB Data  20240514~~ ==================================
# 업로드 From excel to DB
@bp.route('/sortexcel', methods = ['GET','POST'])
def excel_sort():

    minus_time = request.args.get('time', type=int, default=2)

    # 오늘,어제, 현재시간(6자리),현재시간(hhmm00 분봉용) 호출 6자리()
    todate,ydate,totime,totime00 = KFunction.date_info(minus_time)  

    file_path = "//DESKTOP-F5S9HG9/공유 with PC/주마등/검증프로그램/보관/"
    file_name = "Stock_Find Max(rev5.18)_" + ydate +".xlsm"

    print(file_name)     

    # Read the Excel file starting from the 7th row and read data from the third sheet
    #df = pd.read_excel(file_path + file_name, header=None, skiprows=6, names=['Column1', 'Column2', 'Column3'], sheet_name=2) # 모드칼럼
    #칼럼 1~3번까지만
    df = pd.read_excel(file_path + file_name, usecols=[0, 1, 2], header=None, skiprows=6, names=['Column1', 'Column2', 'Column3'], sheet_name=2)

    #print(df)
    # 중복된 항목은 20240516 기준으로 하나만 남김
    #filtered_df = df[(df['Column3'] == 20240516) & (df['Column1'].duplicated() == False)]

    print(df)

    # Find the maximum date value in Column3
    max_date = df['Column3'].max()

    # Filter the DataFrame to include only values with the maximum date value
    filtered_df = df[~df['Column1'].isin(df[df['Column3'] < max_date]['Column1'])]

    # Further exclude duplicates based on the values in Column1
    filtered_df = filtered_df.drop_duplicates(subset='Column1')
    print(filtered_df)
    for index, row in filtered_df.iterrows():
        daily_data_fromvb = DailyDataFromVB(
            # stockcode 값에서 앞에 'A'가 있는 경우 제거하고 사용
            stockcode = row['Column1'].lstrip('A'),
            stockname=row['Column2'],
            stockdate=row['Column3'],
            fromvb = '1'
        )

        # 데이터베이스에 저장
        db.session.add(daily_data_fromvb)
        db.session.commit()
 
    return redirect(url_for('kinvestor._dailylist'))


################################################################################################

#===================== SortExcel daily VB Data  20240514~~ ==================================

#################################################################################################
# 종목 코드 조회   
@bp.route('/dailylist/')
def _dailylist():
    # 입력 파라미터
    page = request.args.get('page', type=int, default=1)  # 페이지
    kw = request.args.get('kw', type=str, default='')
    so = request.args.get('so', type=str, default='recent')
    sdate = request.args.get('stockdate', type=str, default='')  # 페이지
    sel1 = request.args.get('selected1', type=str, default='')
    sel2 = request.args.get('selected2', type=str, default='')
    selectdate = request.args.get('selectdate', type=str, default='')
     

    print(sdate,sel1,sel2,selectdate)
    # 토큰 DB로부터 호출
    static_token = call_token_mock()
    
    # 모든 항목 호출
    # stock_codes = db.session.query(DailyDataFromVB.stockcode, DailyDataFromVB.stockdate).order_by(
    #     desc(DailyDataFromVB.stockdate)).all()

    # 현재 날짜와 시간을 datetime 객체로 얻어옵니다.
    today_date = datetime.now()

    # datetime 객체를 원하는 형식의 문자열로 변환합니다.
    today_date_str = today_date.strftime("%YY%m%d")

    # 날짜 범위 설정
    start_date = '2024-05-16'.replace('-', '')
    end_date = today_date_str

    # 특정 날짜 범위 내의 stock_codes 쿼리
    stock_codes = db.session.query(DailyDataFromVB.stockcode, DailyDataFromVB.stockdate).filter(
        and_(
            DailyDataFromVB.stockdate >= start_date,
            DailyDataFromVB.stockdate <= end_date
        )
    ).order_by((DailyDataFromVB.no)).all()
        
    results = []
    for stockcode, stockdate in stock_codes:
        # Subquery
        subquery = (
            db.session.query(
                ThemeCode.stockcode,
                ThemeCode.themecode,
                ThemeCode.themename
            )
            .filter(ThemeCode.stockcode == stockcode)
            .subquery()
        )

        # Main query
        query = (
            db.session.query(
                DailyDataFromVB.no,
                DailyDataFromVB.stockcode,
                DailyDataFromVB.stockname,
                DailyDataFromVB.stockdate,
                DailyDataFromVB.fromvb,
                DailyDataFromVB.selected1,
                DailyDataFromVB.selected2,
                DailyDataFromVB.selected3,
                DailyDataFromVB.selected4,
                DailyDataFromVB.selecteddate,
                func.group_concat(subquery.c.themecode).label('themecodes'),
                func.group_concat(subquery.c.themename).label('themenames')
            )
            .join(subquery, DailyDataFromVB.stockcode == subquery.c.stockcode)
            .group_by(DailyDataFromVB.stockcode, DailyDataFromVB.stockname)
            .order_by(DailyDataFromVB.no.desc())  # Order by no in descending order
        )

        # Execute the query and fetch the results
        query_result = query.all()
        for row in query_result:
            results.append({
                'no': row.no,
                'stockcode': row.stockcode,
                'stockname': row.stockname,
                'stockdate': row.stockdate,
                'fromvb' : row.fromvb,
                'selecteddate' : row.selecteddate,
                'selected1': row.selected1,
                'selected2' : row.selected2,
                'selected3': row.selected3,
                'selected4': row.selected4,
                'themecodes': row.themecodes,
                'themenames': row.themenames
            })

    #테마 빈도수 조사 
    # 결과를 처리하여 빈도수 계산
    themecode_counter = Counter()
    themename_counter = Counter()

    for result in results:
        # themecodes와 themenames를 콤마로 분리
        if result['themecodes']:
            themecodes = result['themecodes'].split(',')
            themecode_counter.update(themecodes)
        if result['themenames']:
            themenames = result['themenames'].split(',')
            themename_counter.update(themenames)

    # 빈도수에 따라 themecodes와 themenames 정렬
    # sorted_themecodes = sorted(themecode_counter.items(), key=lambda x: x[1], reverse=True)
    # sorted_themenames = sorted(themename_counter.items(), key=lambda x: x[1], reverse=True)
            
    # # 정렬된 결과를 원하는 형식으로 변환
    # sorted_themecodes_str = ', '.join([f"{code} ({count})" for code, count in sorted_themecodes])
    # sorted_themenames_str = ', '.join([f"{name} ({count})" for name, count in sorted_themenames])

    # print("빈도수에 따른 정렬된 Themecodes:", sorted_themecodes_str)
    # print("빈도수에 따른 정렬된 Themenames:", sorted_themenames_str)


    # 상위 20개의 빈도수에 따라 themecodes와 themenames 정렬
    top_themecodes = themecode_counter.most_common(20)
    top_themenames = themename_counter.most_common(20)

    # 결과를 딕셔너리 리스트로 변환
    top_items = []
    for (code, code_count), (name, name_count) in zip(top_themecodes, top_themenames):
        top_items.append({
            'themecode': code,
            'themename': name,
            'count': min(code_count, name_count)  # 두 빈도수 중 최소값을 사용
        })


    print(top_items)
    # 필터링
#     if sdate:
#         results = [result for result in results if result['stockdate'] == sdate]
#     if sel1:
#         results = [result for result in results if result['selected1'] == sel1]
#     if sel2:
#         results = [result for result in results if result['selected2'] == sel2]
#     if selectdate:
#         results = [result for result in results if result['selecteddate'] == selectdate]

#     # 중복 조건을 모두 만족하는 항목만 남김
#     final_results = []
#     for result in results:
#         if (not stockdate or result['stockdate'] == sdate) and \
#         (not sel1 or result['selected1'] == sel1) and \
#         (not sel2 or result['selected2'] == sel2) and \
#         (not selectdate or result['selecteddate'] == selectdate):
#             final_results.append(result)

# # final_results를 클라이언트에 반환

#     # 조회
#     if kw:
#         search = f'%{kw}%'
#         stock_list = [result for result in results if kw.lower() in result['stockcode'].lower() or kw.lower() in result['stockname'].lower()]
#     else:
#         stock_list = results

#     # 페이징
#     per_page = 10
#     offset = (page - 1) * per_page
#     total = len(stock_list)
#     items = stock_list[offset:offset + per_page]

#     pagination = {
#         'page': page,
#         'per_page': per_page,
#         'total': total,
#         'items': items,
#         'has_prev': page > 1,
#         'has_next': offset + per_page < total,
#         'prev_num': page - 1 if page > 1 else None,
#         'next_num': page + 1 if offset + per_page < total else None
#     }

    # 함수 호출
    filtered_results = get_filtered_results(results, sdate, sel1, sel2, selectdate, kw, page)
   # top_theme20 = get_top_theme20()  # 예시로 추가한 함수 호출
    print('스탁데이트는: ', sdate)
    return render_template('stockdata/daily_list.html', 
                           stock_list=filtered_results,
                           sel1 = sel1,
                           sel2 = sel2,
                           selectdate = selectdate,
                           sdate = sdate, 
                           top_theme20=top_items, 
                           page=page, 
                           kw=kw)

# 검색 별도 구분 
def get_filtered_results(results, sdate, sel1, sel2, selectdate, kw, page):
    # 날짜 및 선택 조건 필터링
    if sdate:
        results = [result for result in results if result['stockdate'] == sdate]
    if sel1:
        results = [result for result in results if result['selected1'] == sel1]
    if sel2:
        results = [result for result in results if result['selected2'] == sel2]
    if selectdate:
        results = [result for result in results if result['selecteddate'] == selectdate]

    # 조회 조건 필터링
    if kw:
        search = f'%{kw}%'
        results = [result for result in results if kw.lower() in result['stockcode'].lower() or kw.lower() in result['stockname'].lower()]

    # 페이징 처리
    per_page = 10
    offset = (page - 1) * per_page
    total = len(results)
    items = results[offset:offset + per_page]

    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'items': items,
        'has_prev': page > 1,
        'has_next': offset + per_page < total,
        'prev_num': page - 1 if page > 1 else None,
        'next_num': page + 1 if offset + per_page < total else None
    }

    return pagination



#주식 일별 차트 정보 자세히
@bp.route('/dailydetail', methods=['GET'])
def daily_detail():
    # daishin api
    ki = KValue()
    

    # 입력 파라미터
    page = request.args.get('page', type=int, default=1)  # 페이지
    kw = request.args.get('kw', type=str, default='')
    so = request.args.get('so', type=str, default='recent')
    selected1 = request.args.get('selected1', type=str, default='2')
    selected2 = request.args.get('selected2', type=str, default='2')
    sel1 = request.args.get('sel1', type=str, default='0')
    sel2 = request.args.get('sel2', type=str, default='0')
    sdate = request.args.get('sdate', type=str, default='0')
    selectdate = request.args.get('selectdate', type=str, default='20000101')
    #stockcode1 = request.args.get('stockcode')
    num = request.args.get('no')


    # 빈 문자열을 기본값으로 대체
    if sel1 == '':
        sel1 = '5'
    if sel2 == '':
        sel2 = '5'
    if sdate == '':
        sdate = '5'
    if selectdate == '':
        selectdate = '10000101'

    # 현재 날짜와 시간을 datetime 객체로 얻어옵니다.
    today_date = datetime.now()

    # datetime 객체를 원하는 형식의 문자열로 변환합니다.
    today_date_str = today_date.strftime("%Y%m%d")

    # 문자열을 datetime 객체로 다시 변환합니다. 필요한 경우에만 사용합니다.
    today_date_dt = datetime.strptime(today_date_str, "%Y%m%d")

    # 현재 날짜에서 하루를 빼어 어제의 날짜를 구합니다.
        
    yesterday_date = today_date_dt - timedelta(days=0)

    # 어제의 날짜를 원하는 형식의 문자열로 변환합니다.
    ydate = yesterday_date.strftime("%Y%m%d")
    print(ydate)


    print(sel1,sdate,sel2,selectdate)
    print("번호: ", num, "SEL2: " , sel2)
    if selected1 == '1' or selected1 =='0':
        db.session.query(DailyDataFromVB).filter_by(no=num).update({DailyDataFromVB.selected1: selected1,DailyDataFromVB.selecteddate: ydate}) 
    elif selected2 == '1' or selected2 =='0':
        db.session.query(DailyDataFromVB).filter_by(no=num).update({DailyDataFromVB.selected2: selected2,DailyDataFromVB.selecteddate: ydate}) 


    while True:
        if sel1 == '5' and sel2 == '5' and sdate == '5' and selectdate == '10000101':
            # 조건이 없는 경우 단순히 no 값으로 조회
            stockcode2 = db.session.query(DailyDataFromVB.stockcode, DailyDataFromVB.selected1).filter(DailyDataFromVB.no == num).first()
       
        else:
            max_no = db.session.query(func.max(DailyDataFromVB.no)).scalar()
            
            conditions = []
            if sel1 !='5':
                conditions.append(DailyDataFromVB.selected1 == sel1)
            if sel2 !='5':
                conditions.append(DailyDataFromVB.selected2 == sel2)
            if sdate !='5':
                conditions.append(DailyDataFromVB.stockdate == sdate)
            if selectdate !='10000101':
                conditions.append(DailyDataFromVB.selecteddate == selectdate)

             # 조건이 있는 경우 동적으로 조건을 추가하여 조회
            query = db.session.query(DailyDataFromVB.stockcode, DailyDataFromVB.selected1,DailyDataFromVB.stockdate,DailyDataFromVB.selected2,DailyDataFromVB.selecteddate,DailyDataFromVB.no).filter(DailyDataFromVB.no >= num)
            query = query.filter(and_(*conditions))
            stockcode2 = query.first()
            print(stockcode2)
            print(sel1,sdate,sel2,selectdate,num)
            

            if stockcode2:  
               num = stockcode2[5]         
            
                
            else:
                # num 값을 증가시키거나 다른 종료 조건을 추가하세요
                num_1 = int(num) + 1 
                num = str(num_1)
                if int(max_no) == num_1:
                   break
        

        stockcode = stockcode2[0]  # 한투는 숫자 6자리만 인식함으로 앞자리 삭제
        print("스탁코드는 : ",stockcode)
        # 테마 코드 전달 (20240522)
        theme_codes = ThemeCode.query.filter_by(stockcode=stockcode).with_entities(ThemeCode.themename, ThemeCode.themecode).all()
        #theme_codes = db.session.query(ThemeCode.themename, ThemeCode.themecode).filter(stockcode=stockcode).all()
        theme_dics= [{'themename': name, 'themecode': code} for name, code in theme_codes]  
        
        print(theme_dics)
                
        if not theme_dics:
            num = int(num) + 1
        else:
            break

   
    

    print("스탁2번: ", stockcode2, stockcode2[1])
    n = request.args.get('n', type=str)  
    date_from = request.args.get('from', type=str)
    date_to = request.args.get('to', type=str)
    selected1 = stockcode2[1]

    static_token_mock = call_token_mock()
    # n,date_from,date_to의 경우 값이 없을 경우, get 방식으로 수신 시 ''로, 파라메터가 아예 없을 경우에는 None으로 받음에 따라,
    #처리가 용이하도록 '' 일 경우 None으로 무조건 변경
    print(n,page,date_from)
    if page == 1 and (n !='' or date_from !='' or date_to !='') :

        if n == '': 
            n = None
        if date_from == '' and date_to =='':
            date_from = None
            date_to = None
        
        if date_from is not None and date_to is not None:
            date_from = date_from.replace('-', '') # "-" 제거
            date_to = date_to.replace('-', '') # "-" 제거

        #print("코드1= ", stockcode, "번호= ", n, "시작일=", date_from, "종료일=", date_to)
        if n is None and date_from is None and date_to is None:
            # 현재 날짜를 'YYYYMMDD' 형식의 문자열로 가져오기
            date_to = datetime.today().strftime('%Y%m%d')

            # date_to를 기준으로 120일 전의 날짜 계산
            date_to_dt = datetime.strptime(date_to, '%Y%m%d')
            date_from_dt = date_to_dt - timedelta(days=120)

            # date_from을 'YYYYMMDD' 형식의 문자열로 변환
            date_from = date_from_dt.strftime('%Y%m%d')
            print(f"date_from: {date_from}, date_to: {date_to}")

            # 초기 요청
            stockinfo,stockcandles = ki.period_price(stockcode, static_token_mock, date_from, date_to)
            # stockcandles_dic = json.loads(scandles)
            # stockcandles = stockcandles_dic["output2"]
            # stockinfo = stockcandles_dic["output1"]

            # dailydatas 초기화
            dailydatas = stockcandles

            # 데이터가 정상적으로 수집되었으면 추가 요청
            if stockcandles:
                # data_to - 1일을 구함
                date_to_dt = date_from_dt - timedelta(days=1)
                date_to = date_to_dt.strftime('%Y%m%d')

                # data_from - 120일을 구함
                date_from_dt = date_to_dt - timedelta(days=120)
                date_from = date_from_dt.strftime('%Y%m%d')
                print(f"date_from: {date_from}, date_to: {date_to}")

                # 다시 한 번 데이터 요청
                stockinfo,stockcandles = ki.period_price(stockcode, static_token_mock, date_from, date_to)
                # stockcandles_dic = json.loads(scandles)
                # stockcandles = stockcandles_dic["output2"]
                # stockinfo = stockcandles_dic["output1"]

                # 추가 데이터를 리스트에 합병
                dailydatas.extend(stockcandles)

        print(dailydatas)
        print("코드2= ", stockcode2[0], "번호= ", n, "시작일=", date_from, "종료일=", date_to)
        
        if not (n or date_from):
            return 'Need to provide "n" or "date_from" argument.', 400

        # # 토큰 가져오기
        # static_token = call_token_mock()

        # # 주식 캔들 데이터 가져오기
        # scandles = ki.period_price(stockcode, static_token, date_from, date_to)
    


         # 토근 DB로 부터 호출
        # token = Token_temp.query.first()
        # static_token = token.token_key
        # static_token = call_token_mock()

        # 주식 캔들 데이터 가져오기
       # print("코드3= ", stockcode, "token= ", static_token, "시작일=", date_from, "종료일=", date_to)
        #stockcandles = Creon.get_chart(stockcode, target='A', unit='D', n=n, date_from=date_from, date_to=date_to)
        # print("스탁코드3 :", stockcode2[0])
        # scandles = ki.period_price(stockcode2[0],static_token,date_from,date_to)
        
        # #json to dictionary
        # stockcandles_dic = json.loads(scandles)
        # stockcandles = stockcandles_dic["output2"]
        # stockinfo = stockcandles_dic["output1"]
        #print(stockinfo)

        #1111111=== output1 ; 종목데이터 데이터 === 
       
        #print(stockcandles)
        #기존 DB값 삭제
        StockinfoTemp.query.delete()
        db.session.commit()
          
          #API LIST의 각각 DICT 값을 분리
        # for data in stockinfo:     
        #     stock_code=  data['stck_shrn_iscd']
        #     stockname= data['hts_kor_isnm']
        #     currentvalue = data['stck_prpr']
        #     highvalue  = data['stck_hgpr'] 
        #     lowvalue = data['stck_lwpr']
        #     beginvalue = data['stck_oprc']
        #     diffrate  = data['prdy_ctrt']#등락률
        #     diffval = data['prdy_vrss'] #전일대비 
        #     trade_val = data['acml_vol'] #현재 거래량
        #     pre_tradeval = data['prdy_vol'] #전일거래량
        #     tvol_vsprevious = data['prdy_vrss_vol'] # 전일대비 거래량
        #     faceval = data['stck_fcam'] #액면가
        #     stockvol = data['lstn_stcn'] #상장주수
        #     capital = data['cpfn'] #자본금
        #     stocksum = data['hts_avls'] #시가총액
        #     per = data['per'] #per
        #     eps = data['eps'] # eps
        #     pbr = data['pbr'] #pbr
        #     debtrate = data['itewhol_loan_rmnd_ratem name'] # 전체 융자 잔고 비율
        
        
        # Chartinfo 객체 생성 및 속성 채우기
        stockinfor = StockinfoTemp(
                stockcode= stockinfo['stck_shrn_iscd'],
                stockname=stockinfo['hts_kor_isnm'],
                currentvalue =stockinfo['stck_prpr'],
                highvalue  =stockinfo['stck_hgpr'] ,
                lowvalue =stockinfo['stck_lwpr'],
                beginvalue =stockinfo['stck_oprc'],
                diffrate  =stockinfo['prdy_ctrt'], #등락률
                diffval =stockinfo['prdy_vrss'], #전일대비 
                tradeval =stockinfo['acml_vol'], #현재 거래량
                pre_tradeval =stockinfo['prdy_vol'], #전일거래량
                tvol_vsprevious =stockinfo['prdy_vrss_vol'], # 전일대비 거래량
                faceval =stockinfo['stck_fcam'], #액면가
                stockvol =stockinfo['lstn_stcn'], #상장주수
                capital =stockinfo['cpfn'], #자본금
                stocksum =stockinfo['hts_avls'], #시가총액
                per =stockinfo['per'], #per
                eps =stockinfo['eps'] ,# eps
                pbr =stockinfo['pbr'], #pbr
                debtrate =stockinfo['itewhol_loan_rmnd_ratem name'] # 전체 융자 잔고 비율
            )


            # stockinfor = StockinfoTemp(
            #     stockcode=stock_code,
            #     stockname=stockname,
            #     currentvalue=currentvalue,
            #     highvalue=highvalue,
            #     lowvalue=lowvalue,
            #     beginvalue=beginvalue,
            #     diffrate=diffrate,
            #     diffval=diffval,
            #     tradeval=trade_val,
            #     pre_tradeval=pre_tradeval,
            #     tvol_vsprevious=tvol_vsprevious,
            #     faceval=faceval,
            #     stockvol=stockvol,
            #     capital=capital,
            #     stocksum=stocksum,
            #     per=per,
            #     eps=eps,
            #     pbr=pbr,
            #     debtrate=debtrate
            # )
        # 데이터베이스에 저장
        db.session.add(stockinfor)
        db.session.commit()

 

        #222222=== output2 ; 일별 차트 데이터 === 

        #기존 DB값 삭제
        Chartinfo.query.delete()
        db.session.commit()

        #API LIST의 각각 DICT 값을 분리
        for data in dailydatas:
            date = data['stck_bsop_date']  # 일자
            beginval = str(int(data['stck_oprc']))  # 시가
            highval = str(int(data['stck_hgpr']))  # 고가
            lowval = str(int(data['stck_lwpr']))  # 저가
            endval = str(int(data['stck_clpr']))  # 종가
            tradeval = str(data['acml_vol'])  # 거래량
            tradevol = str(data['acml_tr_pbmn'])  # 거래량

            #traderate  = str(round(data['prdy_vrss_vol_rate'], 1))  # 전일대비 비율
            changeval = str(int(data['prdy_vrss']))  # 전일대비 변동가
            csign = data['prdy_vrss_sign'] # 사인
            if csign =='2':
                changesign = "+"
            elif csign =='5':
                changesign ="-"
            else:
                changesign ="0"
            # 등락률이 지원이 안됨에 따라, 종가에서 변화 계산
            changevol = str(round(int(changeval) /( int(endval) - int(changeval)) *100,1)) + "%"
            #changevol = str(round(float(data['prtt_rate']), 1))  # 등락률
            #foreignbuy = str(data['frgn_ntby_qty']
  



        # Chartinfo 객체 생성 및 속성 채우기
            chartinfo = Chartinfo(
                stockcode=stockcode2[0],
                date=date,
                beginval=beginval,
                highval=highval,
                lowval=lowval,
                endval=endval,
                changeval=changeval,
                tradeval=tradeval,
                tradevol= tradevol,
                changesign=changesign,
                changevol=changevol,
         
            )
        # 데이터베이스에 저장
            db.session.add(chartinfo)
            db.session.commit()
    
    #분봉용 임시 db 생성 ( 향후 한번에 테이블 삭제 필요)
    KTempTable.drop_min_table('all','daily')          
    stock_code = stockcode.replace('A', '')
    KTempTable.create_min_table(stock_code,'daily')  

    ############################# 함수로 변경 필요 ###############################################
    # 분봉데이터 db에 입력
    c_data = []
    # 오늘,어제, 현재시간(6자리),현재시간(hhmm00 분봉용) 호출 6자리()
    todate,ydate,totime,totime00 = KFunction.date_info(0)  
    if totime00 >= '090000':
           
            todate, ydate, totime, totime00 = KFunction.date_info(0)  
            
            # 테이블 이름 구성
            temp_min_table_name = 'daily_min_' + stock_code
            metadata = MetaData()
            table = Table(temp_min_table_name, metadata, autoload_with=db.engine)
            
            data_min_list = []
            if totime00 < '093000':
                data_min = KValue.mindata(stock_code, static_token_mock, totime00)
                print("totime001:", totime00)
                data_min_list.extend(data_min)
            elif totime00 >= '093000' and totime00 <= '235000':
                totime, totime00, before_times = KFunction.beforetime_info(30)
                min_data = db.session.query(table.c.tradetime).filter_by(tradetime=before_times[0]).all()
                print("before_time: ", before_times[0])
                if min_data:
                    data_min = KValue.mindata(stock_code, static_token_mock, totime00)
                    data_min_list.extend(data_min)
                else:
                    totime, totime00, before_times = KFunction.beforetime_info(0)
                    print("시간데이터:", before_times)
                    for beforetime00 in before_times:
                        data_min = KValue.mindata(stock_code, static_token_mock, beforetime00)
                        data_min_list.extend(data_min)
            
            print("min데이터:", data_min_list)

            for min_data in data_min_list:
                record = db.session.query(table).filter_by(tradetime=min_data['stck_cntg_hour']).first()
                
                if record:
                    record.open = min_data['stck_oprc']
                    record.high = min_data['stck_hgpr']
                    record.low = min_data['stck_lwpr']
                    record.close = min_data['stck_prpr']
                    record.tradevol = min_data['cntg_vol']
                else:
                    insert_statement = table.insert().values(
                        tradetime=min_data['stck_cntg_hour'],
                        open=min_data['stck_oprc'],
                        high=min_data['stck_hgpr'],
                        low=min_data['stck_lwpr'],
                        close=min_data['stck_prpr'],
                        tradevol=min_data['cntg_vol']
                    )
                    db.session.execute(insert_statement)

                db.session.commit()
  
  ########################## 함수 제작 필요 ###################################################################


    chart_list = Chartinfo.query.order_by(Chartinfo.date.desc())
    chart_min_list = Chartinfo.query.order_by(Chartinfo.date.desc())
    stock_info = StockinfoTemp.query.all()
    
    # 분봉을 위한 테이블 이름 구성
    temp_min_table_name = 'daily_min_' + stockcode
    metadata = MetaData()
    table = Table(temp_min_table_name, metadata, autoload_with=db.engine)
    #저장된 분봉 데이터 요청
    min_datas = db.session.query(table).order_by(table.c.tradetime.desc())
  
    gragh_list = []
    for chart_info in chart_list:
        row = [
            chart_info.date,
            int(chart_info.beginval),
            int(chart_info.highval),
            int(chart_info.lowval),
            int(chart_info.endval),
            int(chart_info.tradeval),
        ]
        gragh_list.append(row)
    print("일데이터:",gragh_list)
    gragh_min_list = []
    for chart_info in min_datas:
        row = [
            chart_info.tradetime,
            int(chart_info.open),
            int(chart_info.high),
            int(chart_info.low),
            int(chart_info.close),
            int(chart_info.tradevol),
        ]
        gragh_min_list.append(row)
    print("민데이터:",gragh_min_list)

    #print(gragh_list)

    #image_url = url_for('kinvestor.get_chart_image', _external=True)
    image_url = get_chart_image(gragh_list,'day')
    image_min_url = get_chart_image(gragh_min_list,'min')
    #print('이미지url',image_url)

    #image_url = url_for('kinvestor.get_chart_image', _external=True)
    image_url = get_chart_image(gragh_list,'day')
    #print('이미지url',image_url)

    chart_count = Chartinfo.query.count() #총수량
    if chart_count >=5:
        avg_5 = db.session.query(func.avg(Chartinfo.endval)).filter(Chartinfo.no.between(1, 5)).scalar()
    if chart_count >=20:
        avg_20 = db.session.query(func.avg(Chartinfo.endval)).filter(Chartinfo.no.between(1, 20)).scalar()
    if chart_count >=60:
        avg_60 = db.session.query(func.avg(Chartinfo.endval)).filter(Chartinfo.no.between(1, 60)).scalar()
    if chart_count >=120:
        avg_120 = db.session.query(func.avg(Chartinfo.endval)).filter(Chartinfo.no.between(1, 120)).scalar()
    if chart_count >=200:
        avg_200 = db.session.query(func.avg(Chartinfo.endval)).filter(Chartinfo.no.between(1, 200)).scalar()

    # 조회
    
    if kw:
        search = '%%{}%%'.format(kw)
        stock_list = db.session.query(Chartinfo).filter(
        (Chartinfo.stockcode.ilike(search)) | # 주식 코드
        (Chartinfo.stockname.ilike(search)) # 주식 이름
        ).distinct()
 # 페이징
    #print(chart_list)
    chart_list =chart_list.paginate(page=page, per_page=20)
    
    return render_template('stockdata/daily_detail.html', 
                           chart_list=chart_list, 
                           image_url =  image_url,
                           image_min_url = image_min_url,
                           gragh_list = gragh_list,
                           stock_info = stock_info[0],
                           theme_dics =  theme_dics,
                           page=page, 
                           kw=kw, 
                           total_pages=chart_count, 
                           stockcode=stockcode,
                           avg_5 = avg_5,
                           avg_20 = avg_20,
                           num = num,
                           selected1 = selected1,
                           selected2 = selected2,
                           sel1 = sel1,
                           sel2= sel2,
                           sdate = sdate,
                           selectdate = selectdate
                        #    avg_60 = avg_60,
                        #    avg_120 = avg_120,
                        #    avg_200 = avg_200
                           )

#========================================================================================#
# 관심종목 20  from vb
#=========================================================================================#
@bp.route('/toobservestock', methods=['GET', 'POST'])
async def to_observe_stock():

    kt = KTrade()
    kv = KValue()
    kr = RealTimeData()
    kd = RealData()

    # 입력파라메터
    page = request.args.get('page', type=int, default=1)  # 페이지
    kw = request.args.get('kw', type=str, default='')
    so = request.args.get('so', type=str, default='recent')
    r_method = request.args.get('r_method', type=str, default='continue')
    if request.method == 'POST':
        print('POST')
    else:
       
        # 현재 날짜와 시간을 datetime 객체로 얻어옵니다.
        today_date = datetime.now()

        # datetime 객체를 원하는 형식의 문자열로 변환합니다.
        today_date_str = today_date.strftime("%Y%m%d")

        # 문자열을 datetime 객체로 다시 변환합니다. 필요한 경우에만 사용합니다.
        today_date_dt = datetime.strptime(today_date_str, "%Y%m%d")

        # 현재 날짜에서 하루를 빼어 어제의 날짜를 구합니다.
        
        yesterday_date = today_date_dt - timedelta(days=0)

        # 어제의 날짜를 원하는 형식의 문자열로 변환합니다.
        ydate = yesterday_date.strftime("%Y%m%d")
        print(ydate)
        trade_type = 'mock'  # mock : 모의투자 , real : 실투자
        # SQLAlchemy ORM을 사용하여 데이터베이스에서 데이터를 가져옵니다.
        #query = StockListFromVB.query.filter_by(stockdate=ydate)
        stockcodes = DailyDataFromVB.query.filter_by(selected1 = '1')


        ObserveStockFromVB.query.delete()
        db.session.commit()
        for que in stockcodes:
            ob_fromvb = ObserveStockFromVB(
                stockcode = que.stockcode.replace('A', ''),
                stockname= que.stockname,
                stockdate= que.stockdate,
                method_1 = que.selected1,
                method_2 = today_date_dt  #매수 날짜
            )
            db.session.add(ob_fromvb)
            db.session.commit()

        query = ObserveStockFromVB.query.filter_by(method_1='1')


        # 페이징 처리
        query = query.paginate(page=page, per_page=10)
        print(query.items)
        static_token = call_token_mock()
        current_value = []
        code_list = []
        for value in query.items:
            stockcode = value.stockcode.replace('A', '')  # 한투는 숫자 6자리만 인식함으로 앞자리 삭제
            print("스탁코드는: ", stockcode)

            # 실호가,실체결가,체결통보
            # code_list = [['1','H0STASP0',stockcode],['1','H0STCNT0',stockcode],['1','H0STCNI9','njsk2002']] # 모의: H0STCNI9 실:H0STCNI0
            c_list = ['1', 'H0STCNT0', stockcode]  # 모의: H0STCNI9 실:H0STCNI0
            code_list.append(c_list)
            # 현재가는 속도가 늦어서 사용 보류
            # c_data = KValue.current_value(stockcode,static_token)
        print(code_list)

        # 페이지가 1보다 큰 경우에는 새로운 종목 코드 리스트를 가져옴
        if page > 1:
            # 현재 페이지에서 가져올 종목 코드 리스트를 생성
            new_code_list = []
            for value in query.items:
                stockcode = value.stockcode.replace('A', '')  # 한투는 숫자 6자리만 인식함으로 앞자리 삭제
                c_list = ['1', 'H0STCNT0', stockcode]  # 모의: H0STCNI9 실:H0STCNI0
                new_code_list.append(c_list)

            # 이전에 실행 중인 connect 메서드를 중지
            await kd.stop()
            #asyncio.kd.get_event_loop().close()
            print("제발스탑 스탑 스탑")

            # 새로운 종목 코드 리스트로 connect 메서드 실행
            #rdata = await asyncio.create_task(kd.connect(new_code_list, trade_type, r_method))
            #print("알데이터는 = ", rdata)
        else:
            # 페이지가 1인 경우에는 이전에 사용한 코드 리스트로 connect 메서드 실행
            #rdata = await asyncio.create_task(kd.connect(code_list, trade_type, r_method))
            await kd.start(code_list, trade_type, r_method)
        current_value = '10'
        time.sleep(100 / 1000)

        if kw:
            search = '%%{}%%'.format(kw)
            stock_list = db.session.query(Stockinfo).filter(
                (Stockinfo.stockcode.ilike(search)) |  # 주식 코드
                (Stockinfo.stockname.ilike(search))  # 주식 이름
            ).distinct()
        # 페이징

        #return render_template('stockdata/observe_list.html', stock_list=rdata, page=page, kw=kw)
        return 'ok'

########### theme 동일 종목 호출 #########################################################################
@bp.route('/themedata/', methods=['POST'])
def theme_data():
        
        ki = KValue()
        # 클라이언트로부터 받은 JSON 데이터 추출
        data = request.get_json()

        # 받은 데이터 확인
        theme_code = data['themeCode']
        theme_name = data['themeName']

        print(theme_code,theme_name)
        
        stockcodes = ThemeCode.query.filter_by(themecode=theme_code).with_entities(ThemeCode.stockcode, ThemeCode.no).all()
        #theme_codes = db.session.query(ThemeCode.themename, ThemeCode.themecode).filter(stockcode=stockcode).all()
        #theme_dics= [{'stockcode': code, 'stockno': no} for code, no in stockcodes]
        
        print('스탁코드', stockcodes[0])
        stock_dics = [] 
   
        for code,no in stockcodes:
           
            date_from = '20230101'
            date_to = datetime.today().strftime('%Y%m%d')
            # 토근 DB로 부터 호출
            static_token = call_token_mock()
            scandles = ki.period_price(code, static_token, date_from, date_to)
            time.sleep(300/1000)
            # ki.period_price() 메서드의 반환값이 없는 경우 처리
            if scandles:
                stockcandles_dic = json.loads(scandles)
                stockcandles = stockcandles_dic["output2"]
                stockinfo = stockcandles_dic["output1"]

                if stockinfo.get('hts_kor_isnm') and stockinfo['hts_kor_isnm'] != '':
                    print(stockinfo['hts_kor_isnm'])
                    stock_dics.append({
                        'stockcode': stockinfo['stck_shrn_iscd'],
                        'stockname': stockinfo['hts_kor_isnm'],
                        'currentvalue': stockinfo['stck_prpr'],
                        'highvalue': stockinfo['stck_hgpr'],
                        'lowvalue': stockinfo['stck_lwpr'],
                        'beginvalue': stockinfo['stck_oprc'],
                        'diffrate': stockinfo['prdy_ctrt'],  # 등락률
                        'diffval': stockinfo['prdy_vrss'],  # 전일대비
                        'tradeval': stockinfo['acml_vol'],  # 현재 거래량
                        'pre_tradeval': stockinfo['prdy_vol'],  # 전일거래량
                        'tvol_vsprevious': stockinfo['prdy_vrss_vol'],  # 전일대비 거래량
                        'faceval': stockinfo['stck_fcam'],  # 액면가
                        'stockvol': stockinfo['lstn_stcn'],  # 상장주수
                        'capital': stockinfo['cpfn'],  # 자본금
                        'stocksum': stockinfo['hts_avls'],  # 시가총액
                        'per': stockinfo['per'],  # per
                        'eps': stockinfo['eps'],  # eps
                        'pbr': stockinfo['pbr']  # pbr
                    })
            else:
                print("No data returned from ki.period_price()")
        return jsonify({'stock_dics': stock_dics})


###########################################################################################################
# 기본 종목코드(거래소,코스탁) 업데이트 시를 제외하고는 사용 금지(수동으로 url 입력)
# 엑셀파일 to DB
@bp.route('/stockcodes/', methods=['GET'])
def handle_stockcodes():
    category_stock = request.args.get('category', type=str) #페이지

    if category_stock == 'kospi':
        file_path = "C:/projects/download/kinvesotor/"
        file_name = "kospi_code.xlsx"

        wb = openpyxl.load_workbook(file_path + file_name)  # 엑셀 파일을 엽니다.
        ws = wb["Sheet1"]  # "Sheet3" 시트를 선택합니다.

        excel_to_list_all = []  # 엑셀 전체 데이터를 담을 리스트를 초기화합니다.

        seen = set()  # 중복 데이터를 확인하기 위한 집합을 초기화합니다.

        for index, row in enumerate(ws.rows):  # 모든 행을 반복합니다.
            if index >= 2:  # 7번째 행부터 데이터를 저장합니다.
                excel_to_list1 = []  # 한 행의 데이터를 담을 리스트를 초기화합니다.
                duplicate_check = (row[0].value, row[2].value)  # 중복 데이터를 확인하기 위한 튜플을 생성합니다.

                if duplicate_check not in seen:  # 중복이 아닌 경우에만 추가합니다.
                    for cell in row:  # 행의 각 셀을 반복합니다.
                        excel_to_list1.append(cell.value)  # 셀의 값을 리스트에 추가합니다.

                    excel_to_list_all.append(excel_to_list1)  # 행의 데이터 리스트를 전체 데이터 리스트에 추가합니다.
                    seen.add(duplicate_check)  # 중복 데이터 집합에 추가합니다.

        print(excel_to_list_all[5][2])  # 전체 데이터 리스트를 출력합니다.

        # Chartinfo 객체 생성 및 속성 채우기

        for i in range(len(excel_to_list_all) - 1):  # excel_to_list_all의 길이만큼 반복합니다.
            stocklist_fromInvestor = StockListFromKInvestor(
                stockcode= excel_to_list_all[i][0],
                stockstdcode= excel_to_list_all[i][1],
                stockname= excel_to_list_all[i][2],
                cate_large= excel_to_list_all[i][5], #업종구분 대
                cate_medium= excel_to_list_all[i][6], #중
                cate_small= excel_to_list_all[i][7], # 소
                x1_stock= excel_to_list_all[i][37], #거래정지
                x2_stock = excel_to_list_all[i][38], #정리매매
                x3_stock = excel_to_list_all[i][39], # 관리종목
                warning_stock = excel_to_list_all[i][40], #시장경고
                pre_warning_stock = excel_to_list_all[i][41], #경고예고
                x4_stock  = excel_to_list_all[i][42],  # 불성실공시
                backdoor_listing = excel_to_list_all[i][43], # 우회상장
                facevalue  = excel_to_list_all[i][51], #액면가
                listingdate = excel_to_list_all[i][52], #상장일
                listingvolume = excel_to_list_all[i][53], #상장주수
                capital = excel_to_list_all[i][54], # 자본금
                closingmonth = excel_to_list_all[i][55], # 결산월
                pre_stock = excel_to_list_all[i][57], #우선주유무
                category = excel_to_list_all[i][61], #코스피,코스닥,기타(n)
                sales = excel_to_list_all[i][62], # 매출액
                sales_profit = excel_to_list_all[i][63],
                ordinary_profit = excel_to_list_all[i][64],
                profit = excel_to_list_all[i][65],
                roe = excel_to_list_all[i][66],
                basedyear = excel_to_list_all[i][67],
                stock_amount = excel_to_list_all[i][68]

                
                )
            
            # 데이터베이스에 저장
            db.session.add(stocklist_fromInvestor)
            db.session.commit()

    if category_stock == 'kosdaq':
            file_path = "C:/projects/download/kinvesotor/"
            file_name = "kosdaq_code.xlsx"

            wb = openpyxl.load_workbook(file_path + file_name)  # 엑셀 파일을 엽니다.
            ws = wb["Sheet1"]  # "Sheet3" 시트를 선택합니다.

            excel_to_list_all = []  # 엑셀 전체 데이터를 담을 리스트를 초기화합니다.

            seen = set()  # 중복 데이터를 확인하기 위한 집합을 초기화합니다.

            for index, row in enumerate(ws.rows):  # 모든 행을 반복합니다.
                if index >= 2:  # 2번째 행부터 데이터를 저장합니다.
                    excel_to_list1 = []  # 한 행의 데이터를 담을 리스트를 초기화합니다.
                    duplicate_check = (row[0].value, row[2].value)  # 중복 데이터를 확인하기 위한 튜플을 생성합니다.

                    if duplicate_check not in seen:  # 중복이 아닌 경우에만 추가합니다.
                        for cell in row:  # 행의 각 셀을 반복합니다.
                            excel_to_list1.append(cell.value)  # 셀의 값을 리스트에 추가합니다.

                        excel_to_list_all.append(excel_to_list1)  # 행의 데이터 리스트를 전체 데이터 리스트에 추가합니다.
                        seen.add(duplicate_check)  # 중복 데이터 집합에 추가합니다.

            print(excel_to_list_all[0][1])  # 전체 데이터 리스트를 출력합니다.

            # Chartinfo 객체 생성 및 속성 채우기

            for i in range(len(excel_to_list_all) - 1):  # excel_to_list_all의 길이만큼 반복합니다.
                stocklist_fromInvestor = StockListFromKInvestor(
                    stockcode= excel_to_list_all[i][0],
                    stockstdcode= excel_to_list_all[i][1],
                    stockname= excel_to_list_all[i][2],
                    cate_large= excel_to_list_all[i][5], #업종구분 대
                    cate_medium= excel_to_list_all[i][6], #중
                    cate_small= excel_to_list_all[i][7], # 소
                    x1_stock= excel_to_list_all[i][32], #거래정지
                    x2_stock = excel_to_list_all[i][33], #정리매매
                    x3_stock = excel_to_list_all[i][34], # 관리종목
                    warning_stock = excel_to_list_all[i][35], #시장경고
                    pre_warning_stock = excel_to_list_all[i][36], #경고예고
                    x4_stock  = excel_to_list_all[i][37],  # 불성실공시
                    backdoor_listing = excel_to_list_all[i][38], # 우회상장
                    facevalue  = excel_to_list_all[i][46], #액면가
                    listingdate = excel_to_list_all[i][47], #상장일
                    listingvolume = excel_to_list_all[i][48], #상장주수
                    capital = excel_to_list_all[i][49], # 자본금
                    closingmonth = excel_to_list_all[i][50], # 결산월
                    pre_stock = excel_to_list_all[i][52], #우선주유무
                    category = excel_to_list_all[i][53], #코스피,코스닥,기타(n)
                    sales = excel_to_list_all[i][56], # 매출액
                    sales_profit = excel_to_list_all[i][57],
                    ordinary_profit = excel_to_list_all[i][58],
                    profit = excel_to_list_all[i][59],
                    roe = excel_to_list_all[i][60],
                    basedyear = excel_to_list_all[i][61],
                    stock_amount = excel_to_list_all[i][62]

                    
                    )
                # 데이터베이스에 저장
                db.session.add(stocklist_fromInvestor)
                db.session.commit()


    if category_stock == 'theme':
            file_path = "C:/projects/download/kinvesotor/"
            file_name = "theme_code.xlsx"

            wb = openpyxl.load_workbook(file_path + file_name)  # 엑셀 파일을 엽니다.
            ws = wb["Sheet1"]  # "Sheet3" 시트를 선택합니다.

            excel_to_list_all = []  # 엑셀 전체 데이터를 담을 리스트를 초기화합니다.

            seen = set()  # 중복 데이터를 확인하기 위한 집합을 초기화합니다.

            for index, row in enumerate(ws.rows):  # 모든 행을 반복합니다.
                if index >= 2:  # 2번째 행부터 데이터를 저장합니다.
                    excel_to_list1 = []  # 한 행의 데이터를 담을 리스트를 초기화합니다.
                    duplicate_check = (row[0].value, row[2].value)  # 중복 데이터를 확인하기 위한 튜플을 생성합니다.

                    if duplicate_check not in seen:  # 중복이 아닌 경우에만 추가합니다.
                        for cell in row:  # 행의 각 셀을 반복합니다.
                            excel_to_list1.append(cell.value)  # 셀의 값을 리스트에 추가합니다.

                        excel_to_list_all.append(excel_to_list1)  # 행의 데이터 리스트를 전체 데이터 리스트에 추가합니다.
                        seen.add(duplicate_check)  # 중복 데이터 집합에 추가합니다.

            print(excel_to_list_all[0][1])  # 전체 데이터 리스트를 출력합니다.

            # Chartinfo 객체 생성 및 속성 채우기

            for i in range(len(excel_to_list_all) - 1):  # excel_to_list_all의 길이만큼 반복합니다.
                theme_code = ThemeCode(
                    themecode= excel_to_list_all[i][0],
                    themename= excel_to_list_all[i][1],
                    stockcode= excel_to_list_all[i][2]           
                    )
                # 데이터베이스에 저장
                db.session.add(theme_code)
                db.session.commit()

    return 'ok'

### 대신증권용 ##### 기본 종목코드(거래소,코스탁) 업데이트 시를 제외하고는 사용 금지(수동으로 url 입력) 
# @bp.route('/stockcodes/', methods=['GET'])
# def handle_stockcodes():
#     c = Creon()
#     c.avoid_reqlimitwarning()
#     market = request.args.get('market')
#     stock_info_list = []
#     if market == 'kospi':
#         stock_info_list = c.get_stockcodes(1)  # StockInfoVO 객체들의 리스트를 반환하는 함수 호출

#         for stock_info in stock_info_list:
#             stockcode = stock_info.stockcode
#             secondcode = stock_info.secondcode
#             stockname = stock_info.stockname
#             currentvalue = stock_info.stdprice
#             print(stockcode)

#             stockinfo = Stockinfo(stockcode=stockcode,category="1", secondcode=secondcode ,stockname=stockname,currentvalue=currentvalue,create_date=datetime.now(),modify_date=datetime.now())
#             db.session.add(stockinfo)
#             db.session.commit()
    
#         return 'Data saved to database successfully'
    
#     elif market == 'kosdaq':
#         stock_info_list = c.get_stockcodes(2)  # StockInfoVO 객체들의 리스트를 반환하는 함수 호출
        
#         for stock_info in stock_info_list:
#             stockcode = stock_info.stockcode
#             secondcode = stock_info.secondcode
#             stockname = stock_info.stockname
#             currentvalue = stock_info.stdprice
#             print(stockcode)

#             stockinfo = Stockinfo(stockcode=stockcode,category="2", secondcode=secondcode ,stockname=stockname,currentvalue=currentvalue,create_date=datetime.now(),modify_date=datetime.now())
#             db.session.add(stockinfo)
#             db.session.commit()
    
#         return 'Data saved to database successfully'
#     else:
#         return '"market" should be one of "kospi" and "kosdaq".', 400


    

# @bp.route('/stockstatus', methods=['GET'])
# def handle_stockstatus():
#     c = Creon()
#     c.avoid_reqlimitwarning()
#    # stockcode = request.args.get('A000020')
#     stockcode = 'A000020'
#     if not stockcode:
#         return '', 400
#     status = c.get_stockstatus(stockcode)
#     return jsonify(status)

# @bp.route('/stockcandles', methods=['GET'])
# def handle_stockcandles():
#     c = Creon()
#     c.avoid_reqlimitwarning()
#     stockcode = request.args.get('code')
#     n = request.args.get('n')
#     #n = 10
#     date_from = request.args.get('from')
#     #date_from = '20240101'
#     date_to = request.args.get('to')
#     #date_to = '20240419'

#     print(date_from)
#     if not (n or date_from):
#         return 'Need to provide "n" or "date_from" argument.', 400
#     stockcandles = c.get_chart(stockcode, target='A', unit='D', n=n, date_from=date_from, date_to=date_to)
#     #print(stockcandles)
#     return jsonify(stockcandles)



# @bp.route('/marketcandles', methods=['GET'])
# def handle_marketcandles():
#     c = Creon()
#     c.avoid_reqlimitwarning()
#     marketcode = request.args.get('code')
#     n = request.args.get('n')
#     date_from = request.args.get('date_from')
#     date_to = request.args.get('date_to')
#     if marketcode == 'kospi':
#         marketcode = '001'
#     elif marketcode == 'kosdaq':
#         marketcode = '201'
#     elif marketcode == 'kospi200':
#         marketcode = '180'
#     else:
#         return [], 400
#     if not (n or date_from):
#         return '', 400
#     marketcandles = c.get_chart(marketcode, target='U', unit='D', n=n, date_from=date_from, date_to=date_to)
#     return jsonify(marketcandles)

# @bp.route('/stockfeatures/', methods=['GET'])
# def handle_stockfeatures():
#     c = Creon()
#     c.avoid_reqlimitwarning()
#     stockcode = request.args.get('code')
#     #stockcode = 'A000020'
#     if not stockcode:
#         return '', 400
#     stockfeatures = c.get_stockfeatures(stockcode)
#     return jsonify(stockfeatures)

# @bp.route('/short', methods=['GET'])
# def handle_short():
#     c = Creon()
#     c.avoid_reqlimitwarning()
#     stockcode = request.args.get('code')
#     n = request.args.get('n')
#     if not stockcode:
#         return '', 400
#     stockfeatures = c.get_shortstockselling(stockcode, n=n)
#     return jsonify(stockfeatures)


